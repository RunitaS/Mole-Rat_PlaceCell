# -*- coding: utf-8 -*-
"""
Single-unit quality metrics for Neuralynx .ntt files.
Outputs a colour-coded Excel workbook with per-unit metrics and individual histograms.

Metrics computed
────────────────
  SNR                  – A_peak / σ_noise  (MAD-based: σ = MAD/0.6745 on first-5-sample baseline)
  ISI violation %      – fraction of ISIs < 2 ms (good: < 1 %)
  Template correlation – mean Pearson r of each spike vs. mean waveform
  Peak/trough time+amp – peak and trough time (ms) and amplitude (µV) on best channel (mean waveform)
  PT time (ms)         – peak-to-trough time on best channel (mean waveform)
  PT ratio             – |peak| / |trough| on best channel (mean waveform)
  Spike duration (ms)  – time from onset of depolarization to end of the after-hyperpolarization
                         on best channel (mean waveform); onset/offset are the baseline threshold
                         crossings that bracket the peak and trough
  Presence ratio       – fraction of 1-s bins that contain ≥ 1 spike
  CSI                  – complex spike index: % of all ISIs in [3, 20] ms
                         where the 2nd spike amplitude < 1st (Bhatt et al. 2020)

References
──────────
Lee D et al. (2018) Exp Neurobiol 27:593-604.
Ludwig KA et al. (2011) J Neural Eng 8:014001.
Bhatt DL et al. / Bhatt et al. 2020 — Cell 2000 (S0092-8674(00)81828-0)
"""

import os
import numpy as np # type: ignore
import pandas as pd # type: ignore
import matplotlib # type: ignore
matplotlib.use('Agg')
import matplotlib.pyplot as plt # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from openpyxl.drawing.image import Image as XLImage # type: ignore

# ── Configuration ──────────────────────────────────────────────────────────────

ROOT_FOLDER  = r'C:/Runita/NMR/analysis/SurgeryPaperSpikeLFP/8477/L_Iso_CSI_Test'
OUTPUT_EXCEL = r'C:/Runita/NMR/analysis/SurgeryPaperSpikeLFP/8477/L_Iso_CSI_Test/QUALITY_METRICS.xlsx'

N_SAMPLES            = 32        # samples per waveform (NTT tetrode standard)
WAVEFORM_DURATION_MS = 5.0       # total duration spanned by the 32-sample waveform snippet
SAMPLE_INTERVAL_MS   = WAVEFORM_DURATION_MS / N_SAMPLES

# ── Waveform plot colours ──────────────────────────────────────────────────────
COLOR_WAVEFORM = '#2C5F8A'
COLOR_PEAK     = '#E06C75'
COLOR_TROUGH   = '#4A90D9'
COLOR_DURATION = '#5A9E6F'

# ── SNR quality thresholds ─────────────────────────────────────────────────────
SNR_LOW_BAD   = 1.5
SNR_LOW_OK    = 2.5
SNR_HIGH_GOOD = 4.0

# ── ISI violation thresholds (%) ──────────────────────────────────────────────
ISI_GOOD_PCT   = 1.0   # < 1 %  → GREEN
ISI_MARGIN_PCT = 5.0   # 1–5 %  → YELLOW, > 5 % → RED

# ── Spike duration onset/offset threshold ──────────────────────────────────────
# Onset (depolarization begin) / offset (hyperpolarization end) are taken as the
# baseline threshold crossings bracketing the trough/peak, where the threshold is
# this many MAD-based baseline sigmas away from the baseline mean.
SPIKE_DURATION_THRESHOLD_K = 2.0

# ── openpyxl fill colours ─────────────────────────────────────────────────────
FILL_RED    = PatternFill(fill_type='solid', fgColor='FFFF9999')
FILL_YELLOW = PatternFill(fill_type='solid', fgColor='FFFFFF99')
FILL_GREEN  = PatternFill(fill_type='solid', fgColor='FF99FF99')
FILL_PURPLE = PatternFill(fill_type='solid', fgColor='FFD9B3FF')
FILL_HEADER = PatternFill(fill_type='solid', fgColor='FF2E4057')

FONT_HEADER = Font(bold=True, color='FFFFFFFF', name='Calibri', size=10)
FONT_BODY   = Font(name='Calibri', size=10)
ALIGN_CTR   = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT  = Alignment(horizontal='left',   vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin'),  right=Side(style='thin'),
    top=Side(style='thin'),   bottom=Side(style='thin'),
)

# ── Neuralynx dtype constants ─────────────────────────────────────────────────
NTT_HEADER_BYTES   = 16 * 1024
NCS_HEADER_BYTES   = 16 * 1024
DEFAULT_ADBITVOLTS = 0.000000195

NTT_DTYPE = np.dtype([
    ('timestamp',   '<u8'),
    ('sc_number',   '<u4'),
    ('cell_number', '<u4'),
    ('params',      '<u4', (8,)),
    ('waveforms',   '<i2', (32, 4)),
])

NCS_RECORD_DTYPE = np.dtype([
    ('timestamp',   '<u8'),
    ('channel_nr',  '<u4'),
    ('sample_freq', '<u4'),
    ('n_valid',     '<u4'),
    ('samples',     '<i2', (512,)),
])

# ── All metric keys (used for consistent NaN rows on error) ───────────────────
ALL_METRIC_KEYS = [
    'n_spikes',
    'peak_to_peak_ch0_uV', 'peak_to_peak_ch1_uV',
    'peak_to_peak_ch2_uV', 'peak_to_peak_ch3_uV',
    'best_channel', 'a_peak_uV', 'sigma_noise_uV', 'snr', 'snr_db',
    'isi_violation_pct', 'template_correlation',
    'pt_best_channel', 'peak_time_ms', 'peak_amp_uV',
    'trough_time_ms', 'trough_amp_uV', 'pt_time_ms', 'pt_ratio',
    'spike_start_time_ms', 'spike_end_time_ms', 'spike_duration_ms',
    'csi',
]

COLUMN_ORDER = ['session', 'unit'] + ALL_METRIC_KEYS


# ── Helpers: file I/O ──────────────────────────────────────────────────────────

def _parse_adbitvolts(path: str) -> float:
    try:
        with open(path, 'rb') as fh:
            header_raw = fh.read(NTT_HEADER_BYTES)
        header_txt = header_raw.decode('latin-1', errors='replace')
        for line in header_txt.splitlines():
            if 'ADBitVolts' in line:
                for p in line.split():
                    try:
                        val = float(p)
                        if 0 < val < 1:
                            return val
                    except ValueError:
                        pass
    except Exception:
        pass
    return DEFAULT_ADBITVOLTS


def load_ntt_waveforms(ntt_path: str):
    adbitvolts = _parse_adbitvolts(ntt_path)
    uv_scale   = adbitvolts * 1e6
    spike_data = np.memmap(ntt_path, dtype=NTT_DTYPE, mode='r',
                           offset=NTT_HEADER_BYTES)
    waveforms  = spike_data['waveforms'].astype(np.float64) * uv_scale  # type: ignore[operator]
    timestamps = spike_data['timestamp'].astype(np.float64)
    return waveforms, timestamps, adbitvolts


# ── SNR: MAD-based noise estimator ───────────────────────────────────────────

def compute_snr_mad(waveforms: np.ndarray) -> dict:
    """
    SNR for tetrode (.ntt) data using the mean-waveform peak amplitude and a
    MAD-based noise estimate from the pre-spike baseline.

    waveforms : (n_spikes, n_samples, n_channels)
    Returns dict with best_channel, a_peak_uV, sigma_noise_uV, snr, snr_db.
    """
    # Transpose to (n_channels, n_samples, n_spikes) for indexing clarity
    wf = np.transpose(waveforms, (2, 1, 0))          # (4, 32, n_spikes)

    mean_wf = wf.mean(axis=2)                         # (4, 32)
    max_idx  = np.unravel_index(np.argmax(np.abs(mean_wf)), mean_wf.shape)
    best_ch  = int(max_idx[0])

    a_peak = float(np.max(np.abs(mean_wf[best_ch, :])))

    baseline = wf[best_ch, :5, :].flatten()           # first 5 samples as noise proxy
    mad = np.median(np.abs(baseline - np.median(baseline)))
    sigma_noise = float(mad / 0.6745)

    if sigma_noise > 0:
        snr    = a_peak / sigma_noise
        snr_db = 20.0 * np.log10(snr) if snr > 0 else float('nan')
    else:
        snr = snr_db = float('nan')

    return {
        'best_channel':   best_ch,
        'a_peak_uV':      round(a_peak,      4),
        'sigma_noise_uV': round(sigma_noise, 4),
        'snr':            round(snr,    4) if not np.isnan(snr)    else float('nan'),
        'snr_db':         round(snr_db, 4) if not np.isnan(snr_db) else float('nan'),
    }


# ── Quality label helpers ─────────────────────────────────────────────────────

def snr_quality(snr_val) -> tuple:
    if snr_val is None or (isinstance(snr_val, float) and np.isnan(snr_val)):
        return ('N/A', PatternFill())
    if snr_val < SNR_LOW_BAD:
        return ('Too low', FILL_RED)
    if snr_val < SNR_LOW_OK:
        return ('Marginal', FILL_YELLOW)
    if snr_val <= SNR_HIGH_GOOD:
        return ('Good', FILL_GREEN)
    return ('Very high', FILL_PURPLE)


def isi_quality(pct) -> tuple:
    if pct is None or (isinstance(pct, float) and np.isnan(pct)):
        return ('N/A', PatternFill())
    if pct < ISI_GOOD_PCT:
        return ('Good', FILL_GREEN)
    if pct < ISI_MARGIN_PCT:
        return ('Marginal', FILL_YELLOW)
    return ('High', FILL_RED)


# ── Individual quality-metric functions ───────────────────────────────────────

def compute_isi_violations(spike_ts_us: np.ndarray, ref_period_us: float = 2000.0) -> float:
    """Return % of ISIs that violate the refractory period (< ref_period_us)."""
    if len(spike_ts_us) < 2:
        return float('nan')
    isis = np.diff(np.sort(spike_ts_us))
    return float(np.sum(isis < ref_period_us) / len(isis) * 100.0)


def compute_template_correlation(waveforms: np.ndarray) -> float:
    """Mean Pearson r between each spike waveform and the mean template, averaged across channels."""
    n_spikes = waveforms.shape[0]
    if n_spikes < 2:
        return float('nan')
    template = waveforms.mean(axis=0)               # (32, 4)
    t = template - template.mean(axis=0)            # demean along time
    w = waveforms - waveforms.mean(axis=1, keepdims=True)  # (n, 32, 4)

    numer    = np.sum(t[np.newaxis] * w, axis=1)   # (n, 4)
    t_sq_sum = np.sum(t ** 2, axis=0)              # (4,)
    w_sq_sum = np.sum(w ** 2, axis=1)              # (n, 4)
    denom    = np.sqrt(t_sq_sum[np.newaxis] * w_sq_sum)

    with np.errstate(invalid='ignore', divide='ignore'):
        corr = np.where(denom > 0, numer / denom, np.nan)

    return float(np.nanmean(corr))


def compute_spike_duration(wf: np.ndarray, peak_idx: int, trough_idx: int,
                            k: float = SPIKE_DURATION_THRESHOLD_K) -> dict:
    """
    Spike duration: time from the onset of depolarization to the end of the
    after-hyperpolarization, on a single channel's mean waveform.

    The baseline (first 5 samples, same convention as the SNR estimator) gives
    a MAD-based sigma. Onset is the last baseline-level sample before the
    earlier of the peak/trough (depolarization begins where the trace first
    departs the baseline band); offset is the first baseline-level sample
    after the later of the peak/trough (hyperpolarization ends where the
    trace returns to the baseline band).

    wf : (32,) mean waveform on the channel used for peak/trough detection.
    """
    n = wf.shape[0]
    baseline      = wf[:5]
    baseline_mean = float(np.mean(baseline))
    mad           = np.median(np.abs(baseline - np.median(baseline)))
    sigma         = float(mad / 0.6745) if mad > 0 else float(np.std(baseline))
    if sigma == 0:
        sigma = float(np.finfo(float).eps)
    threshold = k * sigma

    first_ext = min(peak_idx, trough_idx)
    last_ext  = max(peak_idx, trough_idx)

    start_idx = 0
    for i in range(first_ext, -1, -1):
        if abs(wf[i] - baseline_mean) < threshold:
            start_idx = i
            break

    end_idx = n - 1
    for i in range(last_ext, n):
        if abs(wf[i] - baseline_mean) < threshold:
            end_idx = i
            break

    duration_ms = (end_idx - start_idx) * SAMPLE_INTERVAL_MS

    return {
        'start_idx':          start_idx,
        'start_time_ms':      round(start_idx * SAMPLE_INTERVAL_MS, 5),
        'end_idx':             end_idx,
        'end_time_ms':         round(end_idx * SAMPLE_INTERVAL_MS, 5),
        'spike_duration_ms':   round(duration_ms, 5),
    }


def compute_pt(waveforms: np.ndarray) -> dict:
    """
    waveforms : (n_spikes, 32, 4) microvolts

    Selects the channel whose MEAN waveform reaches the highest peak
    (max amplitude, not peak-to-peak spread), then measures peak/trough
    time, amplitude ratio, and spike duration on that channel's mean
    waveform.
    """
    mean_wf = waveforms.mean(axis=0)          # (32, 4)

    peak_per_channel = mean_wf.max(axis=0)    # (4,)
    best_ch = int(np.argmax(peak_per_channel))
    wf = mean_wf[:, best_ch]                  # (32,)

    peak_idx   = int(np.argmax(wf))
    trough_idx = int(np.argmin(wf))
    peak_val   = float(wf[peak_idx])
    trough_val = float(wf[trough_idx])

    pt_time_ms = abs(peak_idx - trough_idx) * SAMPLE_INTERVAL_MS
    pt_ratio   = abs(peak_val) / abs(trough_val) if trough_val != 0 else float('nan')

    duration_result = compute_spike_duration(wf, peak_idx, trough_idx)

    return {
        'best_channel':   best_ch,
        'peak_idx':       peak_idx,
        'peak_time_ms':   round(peak_idx * SAMPLE_INTERVAL_MS, 5),
        'peak_amp_uV':    round(peak_val, 4),
        'trough_idx':     trough_idx,
        'trough_time_ms': round(trough_idx * SAMPLE_INTERVAL_MS, 5),
        'trough_amp_uV':  round(trough_val, 4),
        'pt_time_ms':     round(pt_time_ms, 5),
        'pt_ratio':       round(pt_ratio, 4) if not np.isnan(pt_ratio) else float('nan'),
        'spike_start_idx':      duration_result['start_idx'],
        'spike_start_time_ms':  duration_result['start_time_ms'],
        'spike_end_idx':        duration_result['end_idx'],
        'spike_end_time_ms':    duration_result['end_time_ms'],
        'spike_duration_ms':    duration_result['spike_duration_ms'],
        'mean_waveform':  wf,   # used for plotting only, dropped before Excel export
    }


# ── Per-file waveform figure ─────────────────────────────────────────────────

def plot_waveform(wf: np.ndarray, result: dict, out_path: str, title: str):
    time_axis = np.arange(N_SAMPLES) * SAMPLE_INTERVAL_MS

    fig, ax = plt.subplots(figsize=(6, 4.2))
    fig.patch.set_facecolor('white')

    ax.plot(time_axis, wf, color=COLOR_WAVEFORM, linewidth=1.8, zorder=3)
    ax.scatter([result['peak_time_ms']], [result['peak_amp_uV']],
               color=COLOR_PEAK, s=60, zorder=5,
               label=f"Peak  {result['peak_amp_uV']:.1f} uV")
    ax.scatter([result['trough_time_ms']], [result['trough_amp_uV']],
               color=COLOR_TROUGH, s=60, zorder=5,
               label=f"Trough  {result['trough_amp_uV']:.1f} uV")

    start_t = result['spike_start_time_ms']
    end_t   = result['spike_end_time_ms']
    ax.scatter([start_t, end_t], [wf[result['spike_start_idx']], wf[result['spike_end_idx']]],
               color=COLOR_DURATION, marker='|', s=220, linewidths=2.2, zorder=6,
               label='Spike start/end')
    ax.axvspan(start_t, end_t, color=COLOR_DURATION, alpha=0.08, zorder=1)

    ax.set_xlabel('Time (ms)', fontsize=10, labelpad=6)
    ax.set_ylabel('Amplitude (uV)', fontsize=10, labelpad=6)
    ax.set_title(title, fontsize=11, fontweight='bold', pad=10)
    ax.spines[['top', 'right']].set_visible(False)
    ax.legend(fontsize=8, framealpha=0.85, loc='best', edgecolor='#CCCCCC')

    note = (f"Ch {result['best_channel']}  |  "
            f"P-T time = {result['pt_time_ms']:.3f} ms  |  "
            f"P/T ratio = {result['pt_ratio']:.3f}  |  "
            f"Spike duration = {result['spike_duration_ms']:.3f} ms")
    ax.text(0.5, -0.22, note, transform=ax.transAxes, fontsize=8, ha='center')

    plt.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)



def compute_csi(waveforms: np.ndarray, spike_ts_us: np.ndarray,
                min_isi_us: float = 3000.0, max_isi_us: float = 20000.0) -> float:
    """
    Complex Spike Index: % of ALL consecutive ISIs that fall in [3, 20] ms
    AND whose second spike has a smaller amplitude than the first.
    (Bhatt et al. 2020 / Ranck 1973)
    """
    if len(spike_ts_us) < 2:
        return float('nan')
    sort_idx  = np.argsort(spike_ts_us)
    ts_sorted = spike_ts_us[sort_idx]
    wf_sorted = waveforms[sort_idx]

    isis     = np.diff(ts_sorted)
    in_range = (isis >= min_isi_us) & (isis <= max_isi_us)

    if not np.any(in_range):
        return 0.0

    mean_pp = (wf_sorted.max(axis=1) - wf_sorted.min(axis=1)).mean(axis=1)  # (n_spikes,)
    first_pp  = mean_pp[:-1][in_range]
    second_pp = mean_pp[1:][in_range]

    n_complex = int(np.sum(second_pp < first_pp))
    return round(float(n_complex / len(isis) * 100.0), 4)


# ── Master metric computation ─────────────────────────────────────────────────

def compute_metrics(ntt_path: str) -> dict:
    waveforms, spike_ts, _ = load_ntt_waveforms(ntt_path)
    n_spikes = waveforms.shape[0]

    nan_row = {k: float('nan') for k in ALL_METRIC_KEYS}
    nan_row['n_spikes'] = 0
    if n_spikes == 0:
        return nan_row

    # ── Per-channel peak-to-peak (mean across spikes) ─────────────────────────
    pp_per_spike        = waveforms.max(axis=1) - waveforms.min(axis=1)  # (n, 4)
    mean_pp_per_channel = pp_per_spike.mean(axis=0)                      # (4,)

    # ── SNR via MAD-based noise estimator ────────────────────────────────────
    snr_metrics = compute_snr_mad(waveforms)

    # ── Waveform shape metrics (peak-to-trough) ───────────────────────────────
    pt_result = compute_pt(waveforms)

    return {
        'n_spikes':             n_spikes,
        'peak_to_peak_ch0_uV':  round(float(mean_pp_per_channel[0]), 4),
        'peak_to_peak_ch1_uV':  round(float(mean_pp_per_channel[1]), 4),
        'peak_to_peak_ch2_uV':  round(float(mean_pp_per_channel[2]), 4),
        'peak_to_peak_ch3_uV':  round(float(mean_pp_per_channel[3]), 4),
        **snr_metrics,
        'isi_violation_pct':    round(compute_isi_violations(spike_ts), 4),
        'template_correlation': round(compute_template_correlation(waveforms), 4),
        'pt_best_channel':      pt_result['best_channel'],
        'peak_time_ms':         pt_result['peak_time_ms'],
        'peak_amp_uV':          pt_result['peak_amp_uV'],
        'trough_time_ms':       pt_result['trough_time_ms'],
        'trough_amp_uV':        pt_result['trough_amp_uV'],
        'pt_time_ms':           pt_result['pt_time_ms'],
        'pt_ratio':             pt_result['pt_ratio'],
        'spike_start_time_ms':  pt_result['spike_start_time_ms'],
        'spike_end_time_ms':    pt_result['spike_end_time_ms'],
        'spike_duration_ms':    pt_result['spike_duration_ms'],
        'csi':                  compute_csi(waveforms, spike_ts),
        '_pt_plot':             pt_result,   # not an Excel column; used for per-unit waveform plots
    }


# ── Batch scan ────────────────────────────────────────────────────────────────

WAVEFORM_PLOT_DIR = os.path.join(os.path.dirname(OUTPUT_EXCEL), 'waveform_plots')
os.makedirs(WAVEFORM_PLOT_DIR, exist_ok=True)

records = []
waveform_plot_entries = []   # [(label, png_path), ...] for Excel embedding
for dirpath, _, filenames in os.walk(ROOT_FOLDER):
    ntt_files = sorted(f for f in filenames if f.lower().endswith('.ntt'))
    if not ntt_files:
        continue
    for ntt_file in ntt_files:
        ntt_path     = os.path.join(dirpath, ntt_file)
        session_name = os.path.relpath(dirpath, ROOT_FOLDER)
        print(f'Processing: {session_name}  |  {ntt_file}')
        try:
            metrics = compute_metrics(ntt_path)
            pt_plot = metrics.pop('_pt_plot', None)
            if pt_plot is not None:
                unit_label = f'{session_name}_{os.path.splitext(ntt_file)[0]}'.replace(os.sep, '_')
                png_path   = os.path.join(WAVEFORM_PLOT_DIR, f'{unit_label}_waveform.png')
                plot_waveform(pt_plot['mean_waveform'], pt_plot, png_path,
                             title=f'{session_name} | {ntt_file}')
                waveform_plot_entries.append((f'{session_name} | {ntt_file}', png_path))
        except Exception as exc:
            print(f'  ERROR: {exc}')
            metrics = {k: None for k in ALL_METRIC_KEYS}
        metrics['session'] = session_name # type: ignore
        metrics['unit']    = ntt_file # type: ignore
        records.append(metrics)

# ── Build DataFrame ───────────────────────────────────────────────────────────

df = pd.DataFrame(records, columns=COLUMN_ORDER)
df.to_excel(OUTPUT_EXCEL, index=False, engine='openpyxl')

# ── Post-process: colour coding via openpyxl ──────────────────────────────────

wb = load_workbook(OUTPUT_EXCEL)
ws = wb.active
assert ws is not None
ws.title = 'Quality Metrics'

COL_WIDTHS = {
    'session':              28, 'unit':                 18,
    'n_spikes':             10,
    'peak_to_peak_ch0_uV':  18, 'peak_to_peak_ch1_uV':  18,
    'peak_to_peak_ch2_uV':  18, 'peak_to_peak_ch3_uV':  18,
    'best_channel':         14, 'a_peak_uV':            14, 'sigma_noise_uV':       16,
    'snr':                  10, 'snr_db':               10,
    'isi_violation_pct':    18, 'template_correlation':  22,
    'pt_best_channel':      14, 'peak_time_ms':          14, 'peak_amp_uV':          14,
    'trough_time_ms':       14, 'trough_amp_uV':         14, 'pt_time_ms':           14,
    'pt_ratio':             12,
    'spike_start_time_ms':  16, 'spike_end_time_ms':     16, 'spike_duration_ms':    16,
    'csi':                  10,
}
for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 14)

for cell in ws[1]:
    cell.fill = FILL_HEADER; cell.font = FONT_HEADER
    cell.alignment = ALIGN_CTR; cell.border = THIN_BORDER
ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

snr_col     = COLUMN_ORDER.index('snr')               + 1
snr_db_col  = COLUMN_ORDER.index('snr_db')            + 1
isi_col     = COLUMN_ORDER.index('isi_violation_pct') + 1

# Two appended quality-label columns (right after the data)
SNR_QUALITY_COL = len(COLUMN_ORDER) + 1
ISI_QUALITY_COL = len(COLUMN_ORDER) + 2

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    # Base styling
    for cell in row:
        cell.font      = FONT_BODY
        cell.border    = THIN_BORDER
        cell.alignment = ALIGN_CTR if cell.column > 1 else ALIGN_LEFT # type: ignore

    # SNR colour
    try:
        snr_val = float(ws.cell(row=row_idx, column=snr_col).value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        snr_val = float('nan')
    snr_lbl, snr_fill = snr_quality(snr_val)
    ws.cell(row=row_idx, column=snr_col).fill    = snr_fill
    ws.cell(row=row_idx, column=snr_db_col).fill = snr_fill

    c = ws.cell(row=row_idx, column=SNR_QUALITY_COL, value=snr_lbl)  # type: ignore[arg-type]
    c.fill = snr_fill; c.font = FONT_BODY; c.alignment = ALIGN_CTR; c.border = THIN_BORDER

    # ISI colour
    try:
        isi_val = float(ws.cell(row=row_idx, column=isi_col).value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        isi_val = float('nan')
    isi_lbl, isi_fill = isi_quality(isi_val)
    ws.cell(row=row_idx, column=isi_col).fill = isi_fill

    c = ws.cell(row=row_idx, column=ISI_QUALITY_COL, value=isi_lbl)  # type: ignore[arg-type]
    c.fill = isi_fill; c.font = FONT_BODY; c.alignment = ALIGN_CTR; c.border = THIN_BORDER

# Quality label column headers
for col_idx, label in [(SNR_QUALITY_COL, 'SNR Quality'), (ISI_QUALITY_COL, 'ISI Quality')]:
    c = ws.cell(row=1, column=col_idx, value=label)
    c.fill = FILL_HEADER; c.font = FONT_HEADER; c.alignment = ALIGN_CTR; c.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = 14

# ── Legend sheet ──────────────────────────────────────────────────────────────

leg = wb.create_sheet('Legend')
leg.column_dimensions['A'].width = 16
leg.column_dimensions['B'].width = 36
leg.column_dimensions['C'].width = 24

legend_rows = [
    ('Colour',  'Meaning',                                  'Threshold'),
    ('GREEN',   'SNR good / ISI low (< 1 %)',               f'SNR {SNR_LOW_OK}–{SNR_HIGH_GOOD}  |  ISI < {ISI_GOOD_PCT} %'),
    ('YELLOW',  'SNR marginal / ISI borderline (1–5 %)',    f'SNR {SNR_LOW_BAD}–{SNR_LOW_OK}  |  ISI {ISI_GOOD_PCT}–{ISI_MARGIN_PCT} %'),
    ('RED',     'SNR too low / ISI high (> 5 %)',           f'SNR < {SNR_LOW_BAD}  |  ISI > {ISI_MARGIN_PCT} %'),
    ('PURPLE',  'SNR very high – check for artefacts',      f'SNR > {SNR_HIGH_GOOD}'),
]
fills_legend = [FILL_HEADER, FILL_GREEN, FILL_YELLOW, FILL_RED, FILL_PURPLE]
for r_idx, (row_data, fill) in enumerate(zip(legend_rows, fills_legend), start=1):
    for c_idx, val in enumerate(row_data, start=1):
        cell = leg.cell(row=r_idx, column=c_idx, value=val)  # type: ignore[arg-type]
        cell.fill = fill
        cell.font = FONT_HEADER if r_idx == 1 else FONT_BODY
        cell.alignment = ALIGN_CTR; cell.border = THIN_BORDER
    leg.row_dimensions[r_idx].height = 20

wb.save(OUTPUT_EXCEL)
print('Excel data saved.')


# ── Histogram helper ──────────────────────────────────────────────────────────

def make_histogram(values, title, xlabel,
                   thresholds=(), threshold_colors=(), threshold_labels=(),
                   bar_color='#4A90D9', n_bins=30):
    vals = np.array([v for v in values if v is not None and not (isinstance(v, float) and np.isnan(v))])

    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor('#F7F9FC')
    ax.set_facecolor('#F7F9FC')

    if len(vals) > 0:
        ax.hist(vals, bins=n_bins, color=bar_color,
                edgecolor='white', linewidth=0.6, zorder=3)

        for xv, col, lbl in zip(thresholds, threshold_colors, threshold_labels):
            ax.axvline(xv, color=col, linewidth=1.5, linestyle='--', zorder=4, label=lbl)

        mean_v   = float(np.mean(vals))
        median_v = float(np.median(vals))
        ax.axvline(mean_v,   color='#E06C75', linewidth=1.8, linestyle='-',  zorder=5,
                   label=f'Mean   {mean_v:.3f}')
        ax.axvline(median_v, color='#2C5F8A', linewidth=1.8, linestyle=':', zorder=5,
                   label=f'Median {median_v:.3f}')

        ax.yaxis.grid(True, color='white', linewidth=0.8, zorder=2)
        ax.set_axisbelow(True)

        summary = (f'n = {len(vals)}\n'
                   f'Mean ± SD\n'
                   f'{mean_v:.3f} ± {float(np.std(vals)):.3f}')
        ax.text(0.02, 0.97, summary, transform=ax.transAxes, fontsize=8.5,
                va='top', ha='left',
                bbox=dict(boxstyle='round,pad=0.4', facecolor='white',
                          edgecolor='#CCCCCC', alpha=0.9))
        ax.legend(fontsize=8.5, framealpha=0.85, loc='upper right', edgecolor='#CCCCCC')
    else:
        ax.text(0.5, 0.5, 'No valid data', transform=ax.transAxes,
                ha='center', va='center', fontsize=12)

    ax.set_xlabel(xlabel, fontsize=11, labelpad=8)
    ax.set_ylabel('Number of units', fontsize=11, labelpad=8)
    ax.set_title(title, fontsize=13, fontweight='bold', pad=12)
    ax.spines[['top', 'right']].set_visible(False)
    plt.tight_layout()
    return fig


# ── Build and save all histograms ─────────────────────────────────────────────

base = os.path.splitext(OUTPUT_EXCEL)[0]

histogram_specs = [
    # (sheet_name,  png_suffix,        df_col,               title,                            xlabel,                          thresholds,                     thr_colors,                               thr_labels,                                              bar_color)
    ('Hist_SNR',     '_hist_SNR',       'snr',                'SNR Distribution',               'SNR  (A_peak / σ_noise)',       (SNR_LOW_BAD, SNR_LOW_OK, SNR_HIGH_GOOD), ('#E06C75','#E5C07B','#C678DD'), (f'Too low {SNR_LOW_BAD}', f'Marginal {SNR_LOW_OK}', f'High {SNR_HIGH_GOOD}'), '#4A90D9'),
    ('Hist_ISI',     '_hist_ISI',       'isi_violation_pct',  'ISI Violation Distribution',     'ISI violations  (%)',           (ISI_GOOD_PCT, ISI_MARGIN_PCT),  ('#98C379','#E5C07B'),                  (f'Good < {ISI_GOOD_PCT}%', f'Marginal < {ISI_MARGIN_PCT}%'),                 '#E06C75'),
    ('Hist_TmplCorr','_hist_TmplCorr',  'template_correlation','Waveform Template Correlation', 'Mean Pearson r  (spike vs template)', (0.90,),                    ('#98C379',),                           ('r = 0.90',),                                                                   '#7CB9E8'),
    ('Hist_PTTime',  '_hist_PTTime',    'pt_time_ms',         'Peak-to-Trough Time',            'PT time  (ms)',                 (),                              (),                                     (),                                                                              '#A8D8A8'),
    ('Hist_SpikeDur','_hist_SpikeDur',  'spike_duration_ms',  'Spike Duration',                 'Spike duration  (ms)',          (),                              (),                                     (),                                                                              '#C9A8E0'),
    ('Hist_PTRatio', '_hist_PTRatio',   'pt_ratio',           'Peak-to-Trough Ratio',           '|peak| / |trough|  (a.u.)',     (1.0,),                          ('#AAAAAA',),                           ('ratio = 1.0',),                                                                '#FFD580'),
    ('Hist_CSI',     '_hist_CSI',       'csi',                'Complex Spike Index (CSI)',       'CSI  (% of ISIs in [3–20 ms] with smaller 2nd spike)', (10.0,), ('#E06C75',),                           ('CSI = 10%',),                                                                  '#98C379'),
]

png_paths = {}
for sheet_name, suffix, col, title, xlabel, thresholds, thr_colors, thr_labels, bar_color in histogram_specs:
    vals = df[col].values if col in df.columns else []
    fig  = make_histogram(vals, title, xlabel,
                          thresholds=thresholds,
                          threshold_colors=thr_colors,
                          threshold_labels=thr_labels,
                          bar_color=bar_color)
    png  = base + suffix + '.png'
    fig.savefig(png, dpi=150, bbox_inches='tight')
    plt.close(fig)
    png_paths[sheet_name] = png
    print(f'Histogram saved: {png}')

# ── Embed all histograms as separate sheets ───────────────────────────────────

wb2 = load_workbook(OUTPUT_EXCEL)

for sheet_name, suffix, col, title, *_ in histogram_specs:
    hs = wb2.create_sheet(sheet_name)
    hs.sheet_view.showGridLines = False
    tc = hs.cell(row=1, column=1, value=title)
    tc.font      = Font(bold=True, size=13, color='FF2E4057', name='Calibri')
    tc.alignment = Alignment(horizontal='left', vertical='center')
    hs.row_dimensions[1].height = 24
    img = XLImage(png_paths[sheet_name])
    img.anchor = 'A3'
    hs.add_image(img)

# ── Embed per-unit waveform plots (one sheet, stacked rows) ───────────────────

if waveform_plot_entries:
    ROWS_PER_PLOT = 34   # label row + ~image height in default-height rows
    wf_sheet = wb2.create_sheet('Waveforms')
    wf_sheet.sheet_view.showGridLines = False
    wf_sheet.column_dimensions['A'].width = 16

    row_cursor = 1
    for label, png_path in waveform_plot_entries:
        tc = wf_sheet.cell(row=row_cursor, column=1, value=label)
        tc.font      = Font(bold=True, size=11, color='FF2E4057', name='Calibri')
        tc.alignment = Alignment(horizontal='left', vertical='center')
        img = XLImage(png_path)
        img.anchor = f'A{row_cursor + 1}'
        wf_sheet.add_image(img)
        row_cursor += ROWS_PER_PLOT

wb2.save(OUTPUT_EXCEL)

# ── Summary ───────────────────────────────────────────────────────────────────

print(f'\nDone. Results saved to {OUTPUT_EXCEL}')
print(f'Total units processed: {len(df)}')

for col, label in [('snr', 'SNR quality'), ('isi_violation_pct', 'ISI quality')]:
    if df[col].notna().any():
        qual_fn = snr_quality if col == 'snr' else isi_quality
        summary = df[col].apply(lambda v: qual_fn(v)[0]).value_counts()
        print(f'\n{label}:')
        for lbl, n in summary.items():
            print(f'  {lbl:12s}: {n}')

for col in ['isi_violation_pct', 'template_correlation', 'pt_time_ms',
            'pt_ratio', 'spike_duration_ms', 'csi']:
    vals = df[col].dropna()
    if len(vals):
        print(f'\n{col}: mean={vals.mean():.4f}  median={vals.median():.4f}  SD={vals.std():.4f}  n={len(vals)}')
