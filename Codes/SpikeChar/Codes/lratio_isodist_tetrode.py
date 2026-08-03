# -*- coding: utf-8 -*-
"""
L-Ratio and Isolation Distance for Neuralynx .ntt tetrode files.
Faithful Python translation of the ADR / Ken Harris mclust MATLAB reference
(IsolationDistance.m, L_Ratio.m from the Redish mclust package).

For each subfolder under ROOT_DIR the script:
  1. Picks the largest .ntt file (most spikes → most complete sort).
  2. Reads cell_number and waveform data (32 samples × 4 channels per spike).
  3. Builds the tetrode feature matrix FD, replicating mclust's default
     ClusterSeparationFeatures = {'feature_Energy', 'feature_WavePC1'}
     (MClustSettings.m) exactly: for each of the 4 channels, one Energy
     value and one WavePC1 (shape-only, correlation-PCA) value, concatenated
     into a single (n_spikes, 8) matrix — the same FD that
     CalculateLRatio_and_IsolationDistance.m builds before calling
     IsolationDistance.m / L_Ratio.m.
  4. For every sorted cluster (cell_number > 0) computes:
       - Isolation Distance  (IsolationDistance.m — Harris et al. 2001)
       - L-ratio             (L_Ratio.m — Schmitzer-Torbert et al. 2005)
  5. Saves a colour-coded Excel file next to the .ntt file.
       Green row  →  L-ratio < 0.2  AND  Isolation Distance > 10
       Plain row  →  does not satisfy both criteria

═══════════════════════════════════════════════════════════════════════════════
ALGORITHM — line-by-line correspondence to the MATLAB reference
═══════════════════════════════════════════════════════════════════════════════

MATLAB mahal(Y, X)
  Returns the SQUARED Mahalanobis distance d² of every row of Y from the
  distribution defined by X (mean and covariance of X, ddof=1).
  Equivalent Python:
      mean  = X.mean(axis=0)
      cov   = np.cov(X.T)          # ddof=1, matches MATLAB
      Ci    = np.linalg.inv(cov)
      diff  = Y - mean
      d_sq  = (diff @ Ci * diff).sum(axis=1)

feature_Energy.m + feature_WavePC1.m (mclust ClusterSeparationFeatures default)
  Per channel:
      Energy   = sqrt(sum(w², samples)) / sqrt(nSamples)
      w_unit   = w / ||w||₂                       # per-spike, unit L2 norm (shape only)
      cv       = cov(w_unit)                      # ddof=1, across spikes
      corr     = cv / (sd ⊗ sd)                    # sd = sqrt(diag(cv))
      pc       = eig(corr), sorted by descending eigenvalue
      wstd     = (w_unit - mean(w_unit)) / sd       # z-score each time sample
      WavePC1  = (wstd @ pc)[:, 0]                  # first PC score only
  FD = [Energy_ch1..4, WavePC1_ch1..4]   (n_spikes, 8)

IsolationDistance.m (Harris et al. 2001)
  ① Undefined if nClusterSpikes > nSpikes/2  → NaN
  ② m        = mahal(FD, FD[ClusterSpikes])   # d² of ALL spikes
  ③ mNoise   = m[OutClu]                      # d² of non-cluster spikes only
  ④ IsoDist  = sort(mNoise)[nClusterSpikes]   # N_C-th smallest d² (NOT sqrt)

L_Ratio.m (Schmitzer-Torbert et al. 2005)
  ① m        = mahal(FD, FD[ClusterSpikes])   # same d² array
  ② df       = size(FD, 2)                    # = 8 (4 channels × [Energy, WavePC1])
  ③ L        = sum(1 - chi2cdf(m[Noise], df)) # = sum(chi2.sf(d², df))
  ④ Lratio   = L / nClusterSpikes

═══════════════════════════════════════════════════════════════════════════════
BUGS FIXED vs the original Python script
═══════════════════════════════════════════════════════════════════════════════

Bug 1 — FD did not match mclust's default feature space
  WRONG : joint or per-channel PCA on raw waveform samples, 3 PCs/channel,
          no Energy term
  RIGHT : FD = feature_Energy + feature_WavePC1 exactly as mclust builds it
          by default (MClustSettings.ClusterSeparationFeatures)
  WHY   : WavePC1 is PCA on unit-L2-normalized waveforms (shape only, energy
          removed) using the *correlation* matrix (equalizes each time
          sample's influence), keeping only PC1 — not raw-sample covariance
          PCA with 3 components. Energy (per-channel L2 norm) carries the
          amplitude information that WavePC1 deliberately discards. Using a
          different feature space changes L-ratio/IsoDist values even though
          the downstream statistics are computed identically.

Bug 2 — Wrong degrees of freedom (df) in L-ratio chi² term
  WRONG : df = n_pcs = 3
  RIGHT : df = size(FD, 2)  (total feature-space dimensionality, now 8)
  WHY   : MATLAB line "df = size(FD,2)" — df is the number of columns of FD.

Bug 3 — Pooled covariance for Mahalanobis distance
  WRONG : cov(all spikes) used to define the cluster's Mahalanobis space
  RIGHT : cov(cluster spikes only) — MATLAB mahal(FD, FD[ClusterSpikes])
  WHY   : Using the pooled covariance inflates the effective cluster radius
          and makes Isolation Distance appear better than it is.

Bug 4 — sqrt taken on Isolation Distance output        ← new vs previous fix
  WRONG : return sqrt(sorted_non_d_sq[N_C - 1])
  RIGHT : return sorted_non_d_sq[N_C - 1]   (raw d², no sqrt)
  WHY   : MATLAB mahal() returns d² and IsolationDistance.m returns that
          value directly without a sqrt: "IsoDist = sorted(nClusterSpikes)".

References
──────────
Harris KD et al. (2001) J Neurophysiol 84:401–14.
Schmitzer-Torbert N et al. (2005) Neuroscience 131:1–11.
Redish AD — mclust (https://github.com/adredish/MClust)
"""

import os
import glob
import numpy as np # type: ignore
from scipy import stats # type: ignore
import openpyxl # type: ignore
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

# ── Configuration ──────────────────────────────────────────────────────────────

ROOT_DIR           = r'C:/Runita/NMR/analysis/SurgeryPaperSpikeLFP/8477/L_Iso_CSI_Test'

# Quality thresholds (for colour-coding only — raw values always saved)
L_RATIO_THRESHOLD  = 0.2
ISO_DIST_THRESHOLD = 10.0

# ── Neuralynx .ntt binary layout ──────────────────────────────────────────────

NTT_HEADER_BYTES = 16 * 1024   # fixed 16 KB ASCII header

NTT_DTYPE = np.dtype([
    ('timestamp',   '<u8'),
    ('sc_number',   '<u4'),
    ('cell_number', '<u4'),
    ('params',      '<u4', (8,)),
    ('waveforms',   '<i2', (32, 4)),   # 32 samples × 4 channels
])

# ── Excel styling ──────────────────────────────────────────────────────────────

FILL_GREEN  = PatternFill(fill_type='solid', fgColor='FF90EE90')
FILL_HEADER = PatternFill(fill_type='solid', fgColor='FF2E4057')
FONT_HEADER = Font(bold=True, color='FFFFFFFF', name='Calibri', size=10)
FONT_BODY   = Font(name='Calibri', size=10)
ALIGN_CTR   = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),  right=Side(style='thin'),
    top=Side(style='thin'),   bottom=Side(style='thin'),
)

# ── NTT reader ────────────────────────────────────────────────────────────────

def load_ntt(path: str):
    """Return (cell_numbers, waveforms) from a Neuralynx .ntt file.

    cell_numbers : (n_spikes,) uint32 — 0 = unsorted / noise
    waveforms    : (n_spikes, 32, 4)  — raw int16 ADC counts
    """
    data = np.memmap(path, dtype=NTT_DTYPE, mode='r', offset=NTT_HEADER_BYTES)
    return data['cell_number'].copy(), data['waveforms'].copy()


# ── Tetrode feature extraction ─────────────────────────────────────────────────

def build_feature_matrix(waveforms: np.ndarray) -> np.ndarray:
    """Construct the FD matrix exactly as mclust's default
    ClusterSeparationFeatures = {'feature_Energy', 'feature_WavePC1'}
    (MClustSettings.m → CalculateFeatures.m).

    Parameters
    ----------
    waveforms : (n_spikes, 32, 4)

    Returns
    -------
    FD : (n_spikes, 8)  float64 — columns
         [Energy_ch1..4, WavePC1_ch1..4], matching the column order
         CalculateFeatures.m produces when looping over
         {'feature_Energy', 'feature_WavePC1'}.

    feature_Energy.m  (one value per channel)
        energy = sqrt(sum(w², samples)) / sqrt(n_samples)

    feature_WavePC1.m  (one value per channel — PC1 only)
        w        = w_raw / ||w_raw||₂  (per-spike unit L2 norm → shape only,
                                         amplitude/energy removed)
        cv       = cov(w)              # ddof=1, across spikes, (samples × samples)
        sd       = sqrt(diag(cv))
        corr     = cv / outer(sd, sd)  # correlation matrix — equalizes each
                                        # time sample's influence, unlike
                                        # covariance-based PCA on raw samples
        pc       = eigenvectors of corr, sorted by descending eigenvalue
        wstd     = (w - mean(w, axis=0)) / sd
        WavePC1  = (wstd @ pc)[:, 0]   # first PC score only (MATLAB iPC = 1)
    """
    _, n_samples, n_channels = waveforms.shape

    energy_blocks  = []
    wavepc1_blocks = []

    for ch in range(n_channels):
        w_raw = waveforms[:, :, ch].astype(np.float64)   # (n_spikes, n_samples)

        # ── feature_Energy.m ──
        energy = np.sqrt(np.sum(w_raw ** 2, axis=1)) / np.sqrt(n_samples)
        energy_blocks.append(energy[:, None])

        # ── feature_WavePC1.m ──
        l2norms = np.sqrt(np.sum(w_raw ** 2, axis=1))
        w = np.divide(w_raw, l2norms[:, None],
                       out=np.zeros_like(w_raw), where=l2norms[:, None] != 0)

        col_mean = w.mean(axis=0)
        col_std  = w.std(axis=0, ddof=1)
        col_std[col_std == 0] = 1e-10   # guard: constant sample column (not in MATLAB)

        cov_w  = np.cov(w, rowvar=False)          # ddof=1, matches MATLAB cov()
        corr_w = cov_w / np.outer(col_std, col_std)

        _, eigvecs = np.linalg.eigh(corr_w)        # ascending eigenvalues
        pc = eigvecs[:, ::-1]                      # descending — matches MATLAB svd(cvn)

        wstd   = (w - col_mean) / col_std
        scores = wstd @ pc
        wavepc1_blocks.append(scores[:, :1])       # PC1 only

    return np.hstack(energy_blocks + wavepc1_blocks)   # (n_spikes, 8)


# ── mahal() — Python equivalent of MATLAB mahal(Y, X) ────────────────────────

def mahal(Y: np.ndarray, X: np.ndarray) -> np.ndarray:
    """Squared Mahalanobis distance of every row of Y from distribution of X.

    Matches MATLAB mahal(Y, X) exactly:
      • Mean and covariance estimated from X (ddof=1, matching MATLAB cov()).
      • d²(i) = (Y[i] - mean_X) @ inv(cov_X) @ (Y[i] - mean_X)
      • Returns d² for all rows of Y (shape (len(Y),)).

    Raises ValueError if X has too few rows to form a non-singular covariance.
    """
    mean_X = X.mean(axis=0)
    cov_X  = np.cov(X.T)             # ddof=1 by default — matches MATLAB
    if cov_X.ndim == 0:              # scalar edge case (1-D feature space)
        cov_X = cov_X.reshape(1, 1)

    # Small regularisation prevents exact singularity from rounding
    cov_X += np.eye(cov_X.shape[0]) * 1e-10

    try:
        cov_inv = np.linalg.inv(cov_X)
    except np.linalg.LinAlgError:
        cov_inv = np.linalg.pinv(cov_X)

    diff = Y - mean_X                # (n_spikes, n_feat)
    d_sq = (diff @ cov_inv * diff).sum(axis=1)   # (n_spikes,)
    return d_sq


# ── IsolationDistance — Python translation of IsolationDistance.m ─────────────

def isolation_distance(FD: np.ndarray, cluster_mask: np.ndarray) -> float:
    """Python translation of IsolationDistance.m (Harris / Redish mclust).

    Parameters
    ----------
    FD           : (n_spikes, n_feat) full feature matrix (all spikes)
    cluster_mask : (n_spikes,) bool — True for spikes in this cluster

    Returns
    -------
    IsoDist : float — the N_C-th smallest squared Mahalanobis distance among
              non-cluster spikes (d², not sqrt), or NaN if undefined.

    MATLAB equivalent
    -----------------
        nClusterSpikes = sum(cluster_mask)
        if nClusterSpikes > nSpikes/2: return NaN
        m       = mahal(FD, FD[ClusterSpikes, :])   # d² of ALL spikes
        mNoise  = m[OutClu]                          # non-cluster subset
        IsoDist = sort(mNoise)[nClusterSpikes]       # N_C-th smallest (1-based)
    """
    n_spikes   = len(FD)
    n_cluster  = int(cluster_mask.sum())

    # Undefined when cluster is the majority (not enough background spikes)
    if n_cluster > n_spikes / 2:
        return np.nan

    cluster_idx    = np.where(cluster_mask)[0]
    non_cluster_idx = np.where(~cluster_mask)[0]

    try:
        m = mahal(FD, FD[cluster_idx])
    except Exception:
        return np.nan

    m_noise = np.sort(m[non_cluster_idx])   # ascending sort of non-cluster d²

    # N_C-th smallest (MATLAB 1-based index nClusterSpikes → Python index n_cluster-1)
    if len(m_noise) < n_cluster:
        return np.nan

    return float(m_noise[n_cluster - 1])    # raw d², NO sqrt — matches MATLAB


# ── L_Ratio — Python translation of L_Ratio.m ────────────────────────────────

def l_ratio(FD: np.ndarray, cluster_mask: np.ndarray) -> float:
    """Python translation of L_Ratio.m (Schmitzer-Torbert / Redish mclust).

    Parameters
    ----------
    FD           : (n_spikes, n_feat) full feature matrix (all spikes)
    cluster_mask : (n_spikes,) bool — True for spikes in this cluster

    Returns
    -------
    Lratio : float, or NaN if cluster is empty.

    MATLAB equivalent
    -----------------
        NoiseSpikes = setdiff(1:nSpikes, ClusterSpikes)   # all non-cluster
        m      = mahal(FD, FD[ClusterSpikes, :])          # d² of ALL spikes
        df     = size(FD, 2)                              # feature dimensionality
        L      = sum(1 - chi2cdf(m[NoiseSpikes], df))     # = sum(chi2.sf(d², df))
        Lratio = L / nClusterSpikes
    """
    n_cluster = int(cluster_mask.sum())
    if n_cluster == 0:
        return np.nan

    cluster_idx     = np.where(cluster_mask)[0]
    non_cluster_idx = np.where(~cluster_mask)[0]

    try:
        m = mahal(FD, FD[cluster_idx])
    except Exception:
        return np.nan

    df      = FD.shape[1]                                # = 8 (4 channels × [Energy, WavePC1])
    m_noise = m[non_cluster_idx]                         # d² of non-cluster spikes
    L       = float(stats.chi2.sf(m_noise, df).sum())   # sum(1 - chi2cdf(d², df))
    return L / n_cluster


# ── Per-folder driver ─────────────────────────────────────────────────────────

def process_folder(folder: str):
    ntt_files = glob.glob(os.path.join(folder, '*.ntt'))
    if not ntt_files:
        return

    ntt_path = max(ntt_files, key=os.path.getsize)
    print(f'  [{os.path.relpath(folder, ROOT_DIR)}]  {os.path.basename(ntt_path)}')

    try:
        cell_numbers, waveforms = load_ntt(ntt_path)
    except Exception as exc:
        print(f'    ERROR reading file: {exc}')
        return

    clusters = np.unique(cell_numbers)
    clusters = clusters[clusters > 0]
    if len(clusters) == 0:
        print('    No sorted clusters found — skipping.')
        return

    # Build FD matrix (all spikes, Energy + WavePC1 per channel — mclust default)
    FD         = build_feature_matrix(waveforms)
    n_features = FD.shape[1]
    print(f'    Feature matrix: {waveforms.shape[0]} spikes × {n_features} dims '
          f'({waveforms.shape[2]} channels × [Energy, WavePC1])')

    rows = []
    for cid in clusters:
        mask = (cell_numbers == cid)
        iso  = isolation_distance(FD, mask)
        lr   = l_ratio(FD, mask)

        status = f'L-ratio={lr:.4f}  IsoD={iso:.4f}' if not (np.isnan(lr) or np.isnan(iso)) \
                 else '[insufficient spikes for covariance estimate]'
        print(f'    Cluster {cid:2d}: n={int(mask.sum()):5d}  {status}')

        rows.append({
            'cluster':  int(cid),
            'n_spikes': int(mask.sum()),
            'l_ratio':  lr,
            'iso_dist': iso,
        })

    _save_excel(rows, ntt_path, folder, n_features)


# ── Excel writer ──────────────────────────────────────────────────────────────

def _fmt(v):
    return round(float(v), 4) if not np.isnan(v) else 'N/A'


def _save_excel(rows: list, ntt_path: str, folder: str, n_features: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cluster Quality'  # type: ignore

    headers    = ['Cluster ID', 'N Spikes',
                  f'L-Ratio\n(< {L_RATIO_THRESHOLD} good)',
                  f'Isolation Distance (d²)\n(> {ISO_DIST_THRESHOLD} good)',
                  'Quality']
    col_widths = [12, 10, 18, 24, 12]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=c, value=h)  # type: ignore
        cell.fill      = FILL_HEADER
        cell.font      = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w  # type: ignore
    ws.row_dimensions[1].height = 36  # type: ignore
    ws.freeze_panes = 'A2'            # type: ignore

    for r, row in enumerate(rows, 2):
        lr  = row['l_ratio']
        iso = row['iso_dist']
        good = (not np.isnan(lr)  and lr  < L_RATIO_THRESHOLD and
                not np.isnan(iso) and iso > ISO_DIST_THRESHOLD)
        fill    = FILL_GREEN if good else PatternFill()
        quality = 'Good' if good else 'Poor'

        for c, val in enumerate(
                [row['cluster'], row['n_spikes'], _fmt(lr), _fmt(iso), quality], 1):
            cell           = ws.cell(row=r, column=c, value=val)  # type: ignore
            cell.fill      = fill
            cell.font      = FONT_BODY
            cell.alignment = ALIGN_CTR
            cell.border    = THIN_BORDER

    # Method footnote
    note_row = len(rows) + 3
    notes = [
        ('Method (Redish mclust / Harris 2001 / Schmitzer-Torbert 2005):', True),
        (f'Feature space: Energy + WavePC1 × 4 channels = '
         f'{n_features} dims (mclust ClusterSeparationFeatures default)', False),
        ('Mahalanobis distance: cluster-only mean & covariance (MATLAB mahal convention)', False),
        ('Isolation Distance reported as d² (squared Mahalanobis), not sqrt — matches mclust', False),
        (f'Thresholds:  L-Ratio < {L_RATIO_THRESHOLD}  AND  '
         f'Isolation Distance > {ISO_DIST_THRESHOLD}  →  GREEN', False),
    ]
    for i, (text, bold) in enumerate(notes):
        cell      = ws.cell(row=note_row + i, column=1, value=text)  # type: ignore
        cell.font = Font(bold=bold, name='Calibri', size=9)
    ws.cell(row=note_row + len(notes) - 1,  # type: ignore
            column=1).fill = FILL_GREEN

    stem     = os.path.splitext(os.path.basename(ntt_path))[0]
    out_path = os.path.join(folder, f'{stem}_LRatio_IsoDist.xlsx')
    wb.save(out_path)
    print(f'    Saved: {os.path.basename(out_path)}')


# ── Main ──────────────────────────────────────────────────────────────────────

def main(root_dir: str = ROOT_DIR):
    print(f'Scanning: {root_dir}\n')
    processed = 0
    for dirpath, _dirs, files in os.walk(root_dir):
        if any(f.lower().endswith('.ntt') for f in files):
            process_folder(dirpath)
            processed += 1
    print(f'\nDone. Processed {processed} folder(s).')


if __name__ == '__main__':
    main()
