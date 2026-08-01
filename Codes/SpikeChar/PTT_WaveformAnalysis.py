# -*- coding: utf-8 -*-
"""
Peak-to-trough waveform analysis for Neuralynx .ntt tetrode files.

For every .ntt file under the specified target directory:
  1. Reads all 32-sample x 4-channel spike waveforms (scaled to microvolts).
  2. Computes the mean waveform per channel and picks the channel whose mean
     waveform reaches the highest peak (max amplitude) -- the "best channel".
  3. On that channel's mean waveform:
       peak   = highest-amplitude sample
       trough = lowest-amplitude sample
       peak-to-trough time  = |peak_sample - trough_sample| * (5 ms / 32 samples)
       peak-to-trough ratio = |peak amplitude| / |trough amplitude|
  4. Saves a per-file waveform figure (best channel, peak/trough marked).

Across all files it also plots:
  - Peak-to-trough time  (per-file bar chart + distribution histogram)
  - Peak-to-trough ratio (per-file bar chart + distribution histogram)

All results are written to a single Excel summary and PNG figures.

Note on timebase
-----------------
Per spec, the entire 32-sample waveform window spans 5 ms (sample interval
= 5/32 = 0.15625 ms), NOT the Neuralynx-default 32 kHz/1 ms window used
elsewhere in this codebase (e.g. spike_quality.py). Adjust
WAVEFORM_DURATION_MS below if a different digitisation window applies.

Assumption
----------
Each .ntt file in the target directory is treated as already containing the spikes of
a single sorted unit (filenames end in "_SS_<id>"), so ALL spikes in a file
are used -- there is no filtering on cell_number.
"""

import os
import glob
import argparse
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Configuration ──────────────────────────────────────────────────────────

N_SAMPLES            = 32
WAVEFORM_DURATION_MS = 5.0
SAMPLE_INTERVAL_MS   = WAVEFORM_DURATION_MS / N_SAMPLES

DEFAULT_ADBITVOLTS = 0.000000195

# ── Neuralynx .ntt binary layout ────────────────────────────────────────────

NTT_HEADER_BYTES = 16 * 1024   # fixed 16 KB ASCII header

NTT_DTYPE = np.dtype([
    ('timestamp',   '<u8'),
    ('sc_number',   '<u4'),
    ('cell_number', '<u4'),
    ('params',      '<u4', (8,)),
    ('waveforms',   '<i2', (32, 4)),   # 32 samples x 4 channels
])

# ── Excel styling ────────────────────────────────────────────────────────────

FILL_HEADER = PatternFill(fill_type='solid', fgColor='FF2E4057')
FONT_HEADER = Font(bold=True, color='FFFFFFFF', name='Calibri', size=10)
FONT_BODY   = Font(name='Calibri', size=10)
ALIGN_CTR   = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT  = Alignment(horizontal='left',   vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),  right=Side(style='thin'),
    top=Side(style='thin'),   bottom=Side(style='thin'),
)

COLOR_WAVEFORM = '#4472C4'
COLOR_PEAK     = '#C00000'
COLOR_TROUGH   = '#2C5F8A'
COLOR_MEAN     = '#E06C75'
COLOR_MEDIAN   = '#264478'

# ── NTT reader ───────────────────────────────────────────────────────────────

def _parse_adbitvolts(path: str) -> float:
    try:
        with open(path, 'rb') as fh:
            header_txt = fh.read(NTT_HEADER_BYTES).decode('latin-1', errors='replace')
        for line in header_txt.splitlines():
            if 'ADBitVolts' in line:
                for p in line.split():
                    try:
                        val = float(p)
                        if 0 < val < 1:
                            return val
                    except ValueError:
                        pass
    except Exception:
        pass
    return DEFAULT_ADBITVOLTS


def load_ntt(path: str) -> np.ndarray:
    """Return waveforms (n_spikes, 32, 4) in microvolts."""
    adbitvolts = _parse_adbitvolts(path)
    uv_scale   = adbitvolts * 1e6
    data = np.memmap(path, dtype=NTT_DTYPE, mode='r', offset=NTT_HEADER_BYTES)
    return data['waveforms'].astype(np.float64) * uv_scale  # type: ignore[operator]


# ── Core peak-to-trough computation ──────────────────────────────────────────

def compute_pt(waveforms: np.ndarray) -> dict:
    """
    waveforms : (n_spikes, 32, 4) microvolts

    Selects the channel whose MEAN waveform reaches the highest peak
    (max amplitude, not peak-to-peak spread), then measures peak/trough
    time and amplitude ratio on that channel's mean waveform.
    """
    mean_wf = waveforms.mean(axis=0)          # (32, 4)

    peak_per_channel = mean_wf.max(axis=0)    # (4,)
    best_ch = int(np.argmax(peak_per_channel))
    wf = mean_wf[:, best_ch]                  # (32,)

    peak_idx   = int(np.argmax(wf))
    trough_idx = int(np.argmin(wf))
    peak_val   = float(wf[peak_idx])
    trough_val = float(wf[trough_idx])

    pt_time_ms = abs(peak_idx - trough_idx) * SAMPLE_INTERVAL_MS
    pt_ratio   = abs(peak_val) / abs(trough_val) if trough_val != 0 else float('nan')

    return {
        'best_channel':   best_ch,
        'peak_idx':       peak_idx,
        'peak_time_ms':   round(peak_idx * SAMPLE_INTERVAL_MS, 5),
        'peak_amp_uV':    round(peak_val, 4),
        'trough_idx':     trough_idx,
        'trough_time_ms': round(trough_idx * SAMPLE_INTERVAL_MS, 5),
        'trough_amp_uV':  round(trough_val, 4),
        'pt_time_ms':     round(pt_time_ms, 5),
        'pt_ratio':       round(pt_ratio, 4) if not np.isnan(pt_ratio) else float('nan'),
        'mean_waveform':  wf,   # used for plotting only, dropped before Excel export
    }


# ── Per-file waveform figure ─────────────────────────────────────────────────

def plot_waveform(wf: np.ndarray, result: dict, out_path: str, title: str):
    time_axis = np.arange(N_SAMPLES) * SAMPLE_INTERVAL_MS

    fig, ax = plt.subplots(figsize=(6, 4.2))
    fig.patch.set_facecolor('white')

    ax.plot(time_axis, wf, color=COLOR_WAVEFORM, linewidth=1.8, zorder=3)
    ax.scatter([result['peak_time_ms']], [result['peak_amp_uV']],
               color=COLOR_PEAK, s=60, zorder=5,
               label=f"Peak  {result['peak_amp_uV']:.1f} uV")
    ax.scatter([result['trough_time_ms']], [result['trough_amp_uV']],
               color=COLOR_TROUGH, s=60, zorder=5,
               label=f"Trough  {result['trough_amp_uV']:.1f} uV")

    ax.set_xlabel('Time (ms)', fontsize=10, labelpad=6)
    ax.set_ylabel('Amplitude (uV)', fontsize=10, labelpad=6)
    ax.set_title(title, fontsize=11, fontweight='bold', pad=10)
    ax.spines[['top', 'right']].set_visible(False)
    ax.legend(fontsize=8, framealpha=0.85, loc='best', edgecolor='#CCCCCC')

    note = (f"Ch {result['best_channel']}  |  "
            f"P-T time = {result['pt_time_ms']:.3f} ms  |  "
            f"P/T ratio = {result['pt_ratio']:.3f}")
    ax.text(0.5, -0.22, note, transform=ax.transAxes, fontsize=8, ha='center')

    plt.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)


# ── Batch processing ──────────────────────────────────────────────────────────

def process_all(root_dir: str, waveform_dir: str) -> pd.DataFrame:
    ntt_files = sorted(glob.glob(os.path.join(root_dir, '**', '*.ntt'), recursive=True))
    print(f'Found {len(ntt_files)} .ntt files under {root_dir}\n')

    os.makedirs(waveform_dir, exist_ok=True)

    records = []
    for path in ntt_files:
        fname = os.path.basename(path)
        stem  = os.path.splitext(fname)[0]
        print(f'Processing: {fname}')
        try:
            waveforms = load_ntt(path)
            n_spikes  = waveforms.shape[0]
            if n_spikes == 0:
                print('  Skipped (0 spikes)')
                continue

            result = compute_pt(waveforms)
            wf     = result.pop('mean_waveform')

            fig_path = os.path.join(waveform_dir, f'{stem}_waveform.png')
            plot_waveform(wf, result, fig_path, fname)

            result['file']     = fname
            result['n_spikes'] = n_spikes
            records.append(result)
        except Exception as exc:
            print(f'  ERROR: {exc}')

    cols = ['file', 'n_spikes', 'best_channel',
            'peak_idx', 'peak_time_ms', 'peak_amp_uV',
            'trough_idx', 'trough_time_ms', 'trough_amp_uV',
            'pt_time_ms', 'pt_ratio']
    return pd.DataFrame(records, columns=cols)


# ── Summary plots: per-file bar + distribution histogram ─────────────────────

def plot_summary(df: pd.DataFrame, col: str, ylabel: str, title: str,
                 out_png: str, bar_color: str):
    sub   = df[df[col].notna()]
    vals  = sub[col].values
    files = sub['file'].values

    fig, axes = plt.subplots(1, 2, figsize=(16, 5.5),
                              gridspec_kw={'width_ratios': [2.2, 1]})
    fig.patch.set_facecolor('white')

    # Left: per-file bar plot
    ax = axes[0]
    ax.set_facecolor('white')
    x = np.arange(len(vals))
    ax.bar(x, vals, color=bar_color, edgecolor='white', linewidth=0.3, zorder=3)
    ax.set_xticks(x)
    ax.set_xticklabels(files, rotation=90, fontsize=4.5)
    ax.set_ylabel(ylabel, fontsize=11, labelpad=8)
    ax.set_title(f'{title} -- per file (n={len(vals)})', fontsize=12,
                 fontweight='bold', pad=10)
    ax.spines[['top', 'right']].set_visible(False)
    ax.yaxis.grid(True, color='#EEEEEE', linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)

    # Right: distribution histogram
    ax2 = axes[1]
    ax2.set_facecolor('white')
    if len(vals):
        ax2.hist(vals, bins=min(30, max(5, len(vals) // 2)), color=bar_color,
                  edgecolor='white', linewidth=0.6, zorder=3)
        mean_v, median_v = float(np.mean(vals)), float(np.median(vals)) # type: ignore
        ax2.axvline(mean_v, color=COLOR_MEAN, linewidth=1.8, linestyle='-',
                    label=f'Mean {mean_v:.3f}')
        ax2.axvline(median_v, color=COLOR_MEDIAN, linewidth=1.8, linestyle=':',
                    label=f'Median {median_v:.3f}')
        ax2.legend(fontsize=8, framealpha=0.85, loc='upper right', edgecolor='#CCCCCC')
    ax2.set_xlabel(ylabel, fontsize=10, labelpad=6)
    ax2.set_ylabel('Number of files', fontsize=10, labelpad=6)
    ax2.set_title('Distribution', fontsize=12, fontweight='bold', pad=10)
    ax2.spines[['top', 'right']].set_visible(False)

    plt.tight_layout()
    fig.savefig(out_png, dpi=200, bbox_inches='tight')
    plt.close(fig)
    print(f'Saved: {out_png}')


# ── Excel writer ───────────────────────────────────────────────────────────────

def save_excel(df: pd.DataFrame, pt_time_png: str, pt_ratio_png: str, out_path: str):
    df.to_excel(out_path, index=False, engine='openpyxl', sheet_name='PTT Metrics')

    wb = load_workbook(out_path)
    ws = wb['PTT Metrics']

    col_widths = {
        'file': 42, 'n_spikes': 10, 'best_channel': 12,
        'peak_idx': 10, 'peak_time_ms': 14, 'peak_amp_uV': 14,
        'trough_idx': 12, 'trough_time_ms': 16, 'trough_amp_uV': 16,
        'pt_time_ms': 14, 'pt_ratio': 12,
    }
    for c_idx, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col, 14)

    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CTR
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CTR if cell.column > 1 else ALIGN_LEFT # type: ignore

    for sheet_name, png_path, title in [
        ('PT Time Plot',  pt_time_png,  'Peak-to-Trough Time'),
        ('PT Ratio Plot', pt_ratio_png, 'Peak-to-Trough Amplitude Ratio'),
    ]:
        hs = wb.create_sheet(sheet_name)
        hs.sheet_view.showGridLines = False
        tc = hs.cell(row=1, column=1, value=title)
        tc.font = Font(bold=True, size=13, color='FF2E4057', name='Calibri')
        hs.row_dimensions[1].height = 24
        img = XLImage(png_path)
        img.anchor = 'A3'
        hs.add_image(img)

    wb.save(out_path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    # Set up argument parsing
    parser = argparse.ArgumentParser(description="Peak-to-trough waveform analysis for .ntt files.")
    parser.add_argument(
        "-d", "--dir", 
        default=os.getcwd(), 
        help="Root directory containing .ntt files (defaults to current working directory)"
    )
    args = parser.parse_args()

    # Generate paths dynamically based on the provided environment/argument
    root_dir = args.dir
    output_dir = os.path.join(root_dir, 'PTT_Analysis')
    waveform_dir = os.path.join(output_dir, 'WaveformFigures')
    output_excel = os.path.join(output_dir, 'PTT_WaveformAnalysis.xlsx')

    # Pass the variables to your processing function
    df = process_all(root_dir, waveform_dir)
    
    if df.empty:
        print('No valid .ntt files processed -- exiting.')
        return

    # Update references to use the local dynamically generated output_dir
    pt_time_png  = os.path.join(output_dir, 'PTT_time_summary.png')
    pt_ratio_png = os.path.join(output_dir, 'PTT_ratio_summary.png')

    plot_summary(df, 'pt_time_ms', 'Peak-to-trough time (ms)',
                 'Peak-to-Trough Time', pt_time_png, '#7CB9E8')
    plot_summary(df, 'pt_ratio', 'Peak-to-trough amplitude ratio',
                 'Peak-to-Trough Ratio', pt_ratio_png, '#FFD580')

    save_excel(df, pt_time_png, pt_ratio_png, output_excel)

    print(f'\nDone. Processed {len(df)} files.')
    print(f'Excel summary: {output_excel}')
    print(f'Per-file waveform figures: {waveform_dir}')
    print(f'\npt_time_ms : mean={df["pt_time_ms"].mean():.4f}  '
          f'median={df["pt_time_ms"].median():.4f}  SD={df["pt_time_ms"].std():.4f}')
    print(f'pt_ratio   : mean={df["pt_ratio"].mean():.4f}  '
          f'median={df["pt_ratio"].median():.4f}  SD={df["pt_ratio"].std():.4f}')

if __name__ == '__main__':
    main()