#!/usr/bin/env python3
"""
Scan .ntt files under INPUT_ROOTS and report, per Day, how many spikes each
tetrode file holds — flagging empty (no-spike) files.

Directory shape expected (as produced by the sorting pipeline):
    <...>\\<Animal>\\<Paradigm>\\Day1\\1_0\\TT1.ntt, TT2.ntt, ...
    <...>\\<Animal>\\<Paradigm>\\Day1\\2_180\\TT1.ntt, TT2.ntt, ...
    <...>\\<Animal>\\<Paradigm>\\Day1\\3_270\\TT1.ntt, TT2.ntt, ...
    <...>\\<Animal>\\<Paradigm>\\Day1\\4_90\\TT1.ntt, TT2.ntt, ...

A "Day" folder is any folder whose immediate subfolders directly contain
.ntt files (the trial folders, e.g. '1_0', '2_180', ...). Each Day folder
becomes one Excel sheet, named after the Day folder (e.g. 'Day1'); trial
folders become columns, naturally sorted by their leading number so
'1_0' < '2_180' < '3_270' < '4_90'.

Column A lists the .ntt filenames found in the first trial folder.
Columns B, C, D, ... hold the spike count for that same filename in each
subsequent trial folder. A file with zero spikes is written as "no spikes"
and the cell is highlighted light red.

A final "Mostly Empty Folders" sheet lists every trial folder (across all
Days) where more than half of its .ntt files are empty, labeled
"<Day>_<trial folder>" (e.g. "Day1_1_0").

Spike count is derived from file size, same NTT layout as concat_ntt.py:
    n_spikes = (filesize - HEADER_BYTES) // RECORD_BYTES
A file at or below HEADER_BYTES has no spike records and counts as empty.

Usage
-----
  python check_ntt_empty_spikes.py
"""

import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ── configuration ──────────────────────────────────────────────────────────
INPUT_ROOTS = [
    r"X:/NMR_group_data/Runita/Data/Ephys_Data/AllSortedData/Tetrode",
]
OUTPUT_XLSX = r"X:/NMR_group_data/Runita/Data/Ephys_Data/AllSortedData/Tetrode/ntt_spike_counts.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

# ── Neuralynx NTT binary layout (see concat_ntt_for_lrat_isodist_v3.py) ──────
HEADER_BYTES = 16384   # ASCII text header, fixed size
RECORD_BYTES = 304     # bytes per spike record
# ─────────────────────────────────────────────────────────────────────────────

DIGIT_CHUNK = re.compile(r'(\d+)')
INVALID_SHEET_CHARS = set(':\\/?*[]')

NO_SPIKES_TEXT = "no spikes"
FILL_NO_SPIKES = PatternFill(fill_type='solid', fgColor='FFFFC7CE')
FONT_NO_SPIKES = Font(color='FF9C0006')
FILL_HEADER = PatternFill(fill_type='solid', fgColor='FF2E4057')
FONT_HEADER = Font(color='FFFFFFFF', bold=True)

MOSTLY_EMPTY_SHEET = "Mostly Empty Folders"
MOSTLY_EMPTY_THRESHOLD = 0.5  # a folder qualifies if more than this fraction of its files are empty


# ── helpers ───────────────────────────────────────────────────────────────────

def natural_sort_key(name: str):
    """Split a name into text/number chunks so '2' sorts before '10'."""
    parts = DIGIT_CHUNK.split(name)
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def count_spikes(filepath: Path) -> int:
    """Number of spike records in an .ntt file (0 if empty/truncated)."""
    size = filepath.stat().st_size
    if size <= HEADER_BYTES:
        return 0
    return (size - HEADER_BYTES) // RECORD_BYTES


def folder_emptiness(folder: Path) -> tuple:
    """(total, empty) counts for .ntt files found directly in folder."""
    total = 0
    empty = 0
    for f in os.listdir(folder):
        if not f.lower().endswith('.ntt'):
            continue
        total += 1
        if count_spikes(folder / f) == 0:
            empty += 1
    return total, empty


def find_day_groups(root: Path):
    """
    Yield (day_dir, trial_dirs) for every folder under root whose immediate
    subfolders directly contain .ntt files. trial_dirs is naturally sorted
    by folder name.
    """
    for dirpath, dirnames, _ in os.walk(root):
        dirpath = Path(dirpath)
        trial_dirs = []
        for dname in dirnames:
            sub = dirpath / dname
            try:
                has_ntt = any(f.lower().endswith('.ntt') for f in os.listdir(sub))
            except OSError:
                has_ntt = False
            if has_ntt:
                trial_dirs.append(sub)
        if trial_dirs:
            trial_dirs.sort(key=lambda p: natural_sort_key(p.name))
            yield dirpath, trial_dirs


def sanitize_sheet_name(name: str) -> str:
    return ''.join(c for c in name if c not in INVALID_SHEET_CHARS)[:31]


def unique_sheet_name(day_dir: Path, used: set) -> str:
    """Prefer the Day folder's own name; disambiguate with parent folders,
    then a numeric suffix, if that collides (e.g. two animals both have Day1)."""
    parts = day_dir.parts
    for n_parts in range(1, min(len(parts), 4) + 1):
        candidate = sanitize_sheet_name("_".join(parts[-n_parts:]))
        if candidate and candidate not in used:
            used.add(candidate)
            return candidate

    base = sanitize_sheet_name(day_dir.name) or "Sheet"
    i = 2
    while True:
        candidate = f"{base[:28]}_{i}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


# ── core logic ────────────────────────────────────────────────────────────────

def write_day_sheet(wb: Workbook, sheet_name: str, day_dir: Path, trial_dirs: list) -> None:
    ws = wb.create_sheet(title=sheet_name)

    headers = ["Filename"] + [t.name for t in trial_dirs]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    first_dir = trial_dirs[0]
    filenames = sorted(
        (f for f in os.listdir(first_dir) if f.lower().endswith('.ntt')),
        key=natural_sort_key,
    )

    print(f"[{sheet_name}]  {len(trial_dirs)} trial folder(s), "
          f"{len(filenames)} file(s) in '{first_dir.name}'")

    for filename in filenames:
        row_values = [filename]
        for trial_dir in trial_dirs:
            filepath = trial_dir / filename
            if not filepath.exists():
                row_values.append("file not found")
                continue
            n_spikes = count_spikes(filepath)
            if n_spikes == 0:
                print(f"    [EMPTY] {trial_dir.name}/{filename}")
                row_values.append(NO_SPIKES_TEXT)
            else:
                row_values.append(n_spikes)
        ws.append(row_values)

        row_idx = ws.max_row
        for col_idx, value in enumerate(row_values[1:], start=2):
            if value == NO_SPIKES_TEXT:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = FILL_NO_SPIKES
                cell.font = FONT_NO_SPIKES

    for col_idx, header in enumerate(headers, start=1):
        width = max(len(header), 14)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def write_summary_sheet(wb: Workbook, mostly_empty: list) -> None:
    """mostly_empty: list of (label, folder_path, empty, total) tuples."""
    ws = wb.create_sheet(title=MOSTLY_EMPTY_SHEET)

    headers = ["Day_Folder", "Full Path", "Empty Files", "Total Files", "% Empty"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    mostly_empty = sorted(mostly_empty, key=lambda t: t[2] / t[3], reverse=True)
    for label, folder_path, empty, total in mostly_empty:
        ws.append([label, str(folder_path), empty, total, empty / total])
        cell = ws.cell(row=ws.max_row, column=5)
        cell.number_format = '0.0%'
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).fill = FILL_NO_SPIKES
            ws.cell(row=ws.max_row, column=col).font = FONT_NO_SPIKES

    widths = [22, 90, 12, 12, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))  # move to front


def process_roots(roots: list, output_xlsx: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    used_names = set()
    n_sheets = 0
    mostly_empty = []  # (label, folder_path, empty, total)

    for root in roots:
        root = Path(root)
        if not root.is_dir():
            print(f"[SKIP] '{root}' is not a directory.")
            continue

        for day_dir, trial_dirs in find_day_groups(root):
            sheet_name = unique_sheet_name(day_dir, used_names)
            write_day_sheet(wb, sheet_name, day_dir, trial_dirs)
            n_sheets += 1

            for trial_dir in trial_dirs:
                total, empty = folder_emptiness(trial_dir)
                if total > 0 and empty / total > MOSTLY_EMPTY_THRESHOLD:
                    label = f"{sheet_name}_{trial_dir.name}"
                    mostly_empty.append((label, trial_dir, empty, total))
                    print(f"    [MOSTLY EMPTY] {label}  ({empty}/{total} files empty)")

    if n_sheets == 0:
        print("No Day folders (with trial subfolders containing .ntt files) were found.")
        return

    if mostly_empty:
        write_summary_sheet(wb, mostly_empty)

    Path(output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    print(f"\nSaved {n_sheets + (1 if mostly_empty else 0)} sheet(s) to {output_xlsx}")


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("Input roots:")
    for r in INPUT_ROOTS:
        print(f"  {r}")
    print(f"Output: {OUTPUT_XLSX}\n{'─' * 60}\n")
    process_roots(INPUT_ROOTS, OUTPUT_XLSX)
    print("Done.")
