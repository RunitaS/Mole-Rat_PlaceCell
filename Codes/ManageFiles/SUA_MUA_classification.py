# -*- coding: utf-8 -*-
"""
Burst frequency range estimation from Neuralynx .ntt files.

For each cluster, builds an ISI histogram and checks whether significant
spike-count power falls in the burst range [ISI_BURST_LO_MS, ISI_BURST_HI_MS].
Clusters that pass the significance threshold have their peak ISI period and
half-maximum limits extracted and saved to an Excel workbook.

Outputs
───────
  BURST_FREQ_RANGE.xlsx   – per-unit table + summary histograms
  Individual ISI PNGs     – one per unit, saved alongside the Excel
"""

import os
import shutil
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Configuration ──────────────────────────────────────────────────────────────
ROOT_FOLDER    = r'X:/NMR_group_data/Runita/Data/Ephys_Data/AllSortedData/Tetrode/Fa8477'
OUTPUT_EXCEL   = r'X:/NMR_group_data/Runita/Data/Ephys_Data/AllSortedData/Tetrode/Fa8477/8477_SUA_MUA_Classification.xlsx'
SUA_SUBDIR_NAME = 'SUA'   # per-folder subdirectory that SUA .ntt files are copied into

NTT_SAMPLE_RATE_HZ = 32000.0

# ── Complex spike filter ──────────────────────────────────────────────────────
CSI_THRESHOLD        = 5.0   # CSI must exceed this (%) for a cell to qualify

# ── SUA / MUA classification ───────────────────────────────────────────────────
RPV_THRESHOLD_MS = 2.0    # absolute refractory period (ms); ISIs below this = violation
RPV_MAX_PCT      = 1    # max refractory-period violation rate (%) to qualify as SUA
WF_SNR_MIN       = 3.0    # min waveform SNR (mean peak-to-peak / 2×baseline noise SD)
WF_CV_MAX        = 0.5    # max waveform amplitude coefficient-of-variation for SUA

# ── Burst ISI range ────────────────────────────────────────────────────────────
ISI_BURST_LO_MS  = 2.0    # lower ISI bound of burst range (ms)
ISI_BURST_HI_MS  = 15.0   # upper ISI bound of burst range (ms)
SIGNIFICANCE_PCT = 5.0    # min % of ISIs in burst range to call it significant
ISI_BIN_WIDTH_MS = 0.5    # histogram bin width (ms)
ISI_HIST_MAX_MS  = 200.0  # histogram x-axis ceiling (ms)

ISI_HIST_BINS_MS = np.arange(0, ISI_HIST_MAX_MS + ISI_BIN_WIDTH_MS, ISI_BIN_WIDTH_MS)

# ── openpyxl styles ───────────────────────────────────────────────────────────
FILL_GREEN  = PatternFill(fill_type='solid', fgColor='FF99FF99')
FILL_YELLOW = PatternFill(fill_type='solid', fgColor='FFFFFF99')
FILL_RED    = PatternFill(fill_type='solid', fgColor='FFFF9999')
FILL_HEADER = PatternFill(fill_type='solid', fgColor='FF2E4057')

FONT_HEADER = Font(bold=True, color='FFFFFFFF', name='Calibri', size=10)
FONT_BODY   = Font(name='Calibri', size=10)
ALIGN_CTR   = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT  = Alignment(horizontal='left',   vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)

# ── Neuralynx dtype constants ─────────────────────────────────────────────────
NTT_HEADER_BYTES   = 16 * 1024
DEFAULT_ADBITVOLTS = 0.000000195

NTT_DTYPE = np.dtype([
    ('timestamp',   '<u8'),
    ('sc_number',   '<u4'),
    ('cell_number', '<u4'),
    ('params',      '<u4', (8,)),
    ('waveforms',   '<i2', (32, 4)),
])

# ── Output column layout ──────────────────────────────────────────────────────
ALL_METRIC_KEYS = [
    'n_spikes',
    'is_sua', 'rpv_rate', 'waveform_snr', 'waveform_cv',
    'csi_og', 'is_complex_cell',
    'burst_power_pct',
    'has_burst_peak',
    'peak_isi_ms',   'peak_freq_hz',
    'lower_isi_ms',  'upper_isi_ms',
    'lower_freq_hz', 'upper_freq_hz',
]
COLUMN_ORDER = ['session', 'unit'] + ALL_METRIC_KEYS


# ── File I/O ──────────────────────────────────────────────────────────────────

def _parse_adbitvolts(path: str) -> float:
    try:
        with open(path, 'rb') as fh:
            header_raw = fh.read(NTT_HEADER_BYTES)
        header_txt = header_raw.decode('latin-1', errors='replace')
        for line in header_txt.splitlines():
            if 'ADBitVolts' in line:
                for p in line.split():
                    try:
                        val = float(p)
                        if 0 < val < 1:
                            return val
                    except ValueError:
                        pass
    except Exception:
        pass
    return DEFAULT_ADBITVOLTS


def load_ntt_waveforms(ntt_path: str):
    adbitvolts = _parse_adbitvolts(ntt_path)
    uv_scale   = adbitvolts * 1e6
    spike_data = np.memmap(ntt_path, dtype=NTT_DTYPE, mode='r',
                           offset=NTT_HEADER_BYTES)
    waveforms  = spike_data['waveforms'].astype(np.float64) * uv_scale  # type: ignore[operator]
    timestamps = spike_data['timestamp'].astype(np.float64)
    return waveforms, timestamps, adbitvolts


# ── Complex Spike Index ───────────────────────────────────────────────────────

def compute_csi(waveforms: np.ndarray, spike_ts_us: np.ndarray,
                min_isi_us: float = 2000.0, max_isi_us: float = 15000.0) -> float:
    """
    Complex Spike Index: % of ALL consecutive ISIs that fall in [2, 15] ms
    AND whose second spike has a smaller amplitude than the first.
    (Bhatt et al. 2020 / Ranck 1973)
    """
    if len(spike_ts_us) < 2:
        return float('nan')
    sort_idx  = np.argsort(spike_ts_us)
    ts_sorted = spike_ts_us[sort_idx]
    wf_sorted = waveforms[sort_idx]

    isis     = np.diff(ts_sorted)
    in_range = (isis >= min_isi_us) & (isis <= max_isi_us)

    if not np.any(in_range):
        return 0.0

    mean_pp   = (wf_sorted.max(axis=1) - wf_sorted.min(axis=1)).mean(axis=1)  # (n_spikes,)
    first_pp  = mean_pp[:-1][in_range]
    second_pp = mean_pp[1:][in_range]

    n_complex = int(np.sum(second_pp < first_pp))
    return round(float(n_complex / len(isis) * 100.0), 4)


# ── SUA / MUA classification ─────────────────────────────────────────────────

def classify_sua_mua(waveforms: np.ndarray, spike_ts_us: np.ndarray) -> dict:
    """
    Classify a cluster as SUA or MUA.  SUA requires ALL three:
      1. RPV rate (% of ISIs < RPV_THRESHOLD_MS) < RPV_MAX_PCT
      2. Waveform SNR (mean peak-to-peak / 2×baseline noise SD) >= WF_SNR_MIN
      3. Waveform amplitude CV <= WF_CV_MAX

    waveforms shape: (n_spikes, 32_samples, 4_wires)
    Baseline noise is estimated from the first 4 pre-trigger samples.
    """
    nan = float('nan')
    result = {'is_sua': False, 'rpv_rate': nan, 'waveform_snr': nan, 'waveform_cv': nan}

    if len(spike_ts_us) < 2:
        return result

    isis_ms  = np.diff(np.sort(spike_ts_us)) / 1000.0
    rpv_rate = float(np.sum(isis_ms < RPV_THRESHOLD_MS) / len(isis_ms) * 100.0)
    result['rpv_rate'] = round(rpv_rate, 3)

    # Mean peak-to-peak amplitude per spike, averaged across the 4 tetrode wires
    pp_per_spike = (waveforms.max(axis=1) - waveforms.min(axis=1)).mean(axis=1)  # (n_spikes,)
    signal       = float(pp_per_spike.mean())

    # Noise SD from the first 4 pre-trigger samples (pre-spike baseline window)
    baseline_sd  = float(waveforms[:, :4, :].std())
    waveform_snr = round(signal / (2.0 * baseline_sd), 3) if baseline_sd > 0 else nan
    waveform_cv  = round(float(pp_per_spike.std() / signal), 3) if signal > 0 else nan

    result['waveform_snr'] = waveform_snr
    result['waveform_cv']  = waveform_cv

    is_sua = (
        rpv_rate < RPV_MAX_PCT
        and not np.isnan(waveform_snr) and waveform_snr >= WF_SNR_MIN
        and not np.isnan(waveform_cv)  and waveform_cv  <= WF_CV_MAX
    )
    result['is_sua'] = bool(is_sua)
    return result


# ── Burst frequency range estimation ─────────────────────────────────────────

def compute_burst_freq_range(spike_ts_us: np.ndarray,
                              lo_ms: float = ISI_BURST_LO_MS,
                              hi_ms: float = ISI_BURST_HI_MS,
                              sig_pct: float = SIGNIFICANCE_PCT,
                              bins_ms: np.ndarray = ISI_HIST_BINS_MS) -> dict:
    """
    Estimate burst frequency range from the ISI histogram in [lo_ms, hi_ms].

    Steps
    ─────
    1. Compute all pairwise consecutive ISIs.
    2. Measure burst_power_pct = % of ISIs in [lo_ms, hi_ms].
    3. If burst_power_pct >= sig_pct, the cluster is flagged as having a
       significant burst peak and the following are extracted:
         peak_isi_ms   – ISI at the histogram maximum within the burst range
         lower_isi_ms  – ISI at the left  half-maximum crossing
         upper_isi_ms  – ISI at the right half-maximum crossing
       Corresponding frequencies (Hz) = 1000 / ISI_ms are also returned.
       Note: lower_freq_hz ↔ upper_isi_ms  (shorter ISI = higher frequency).

    Returns dict with all keys in ALL_METRIC_KEYS except n_spikes.
    """
    nan = float('nan')
    result = {
        'burst_power_pct': nan,
        'has_burst_peak':  False,
        'peak_isi_ms':     nan,  'peak_freq_hz':   nan,
        'lower_isi_ms':    nan,  'upper_isi_ms':   nan,
        'lower_freq_hz':   nan,  'upper_freq_hz':  nan,
    }

    if len(spike_ts_us) < 2:
        return result

    isis_ms = np.diff(np.sort(spike_ts_us)) / 1000.0   # µs → ms

    # ── Burst power: fraction of ISIs inside the burst range ─────────────────
    in_burst = (isis_ms >= lo_ms) & (isis_ms <= hi_ms)
    burst_power_pct = float(np.sum(in_burst) / len(isis_ms) * 100.0)
    result['burst_power_pct'] = round(burst_power_pct, 3)

    if burst_power_pct < sig_pct:
        return result   # not significant – leave peak/limit fields as NaN

    result['has_burst_peak'] = True

    # ── ISI histogram (full range; analyse only burst window) ─────────────────
    counts, edges = np.histogram(isis_ms, bins=bins_ms)
    bin_centers   = 0.5 * (edges[:-1] + edges[1:])

    in_range       = (bin_centers >= lo_ms) & (bin_centers <= hi_ms)
    range_counts   = counts[in_range].astype(float)
    range_centers  = bin_centers[in_range]

    if len(range_counts) == 0 or range_counts.max() == 0:
        return result

    # ── Peak period ───────────────────────────────────────────────────────────
    peak_local = int(np.argmax(range_counts))
    peak_isi   = float(range_centers[peak_local])
    peak_count = float(range_counts[peak_local])
    half_max   = peak_count / 2.0

    # ── Half-maximum lower limit (search left of peak) ────────────────────────
    lower_isi = lo_ms   # fallback: burst range edge
    for i in range(peak_local, -1, -1):
        if range_counts[i] <= half_max:
            if i < peak_local:
                # linear interpolation between bin i and i+1
                x0, x1 = range_centers[i], range_centers[i + 1]
                y0, y1 = range_counts[i],  range_counts[i + 1]
                if y1 != y0:
                    lower_isi = float(x0 + (half_max - y0) / (y1 - y0) * (x1 - x0))
                else:
                    lower_isi = float(x0)
            else:
                lower_isi = float(range_centers[i])
            break

    # ── Half-maximum upper limit (search right of peak) ──────────────────────
    upper_isi = hi_ms   # fallback: burst range edge
    for i in range(peak_local, len(range_counts)):
        if range_counts[i] <= half_max:
            if i > peak_local:
                x0, x1 = range_centers[i - 1], range_centers[i]
                y0, y1 = range_counts[i - 1],  range_counts[i]
                if y0 != y1:
                    upper_isi = float(x0 + (half_max - y0) / (y1 - y0) * (x1 - x0))
                else:
                    upper_isi = float(x0)
            else:
                upper_isi = float(range_centers[i])
            break
 
    # Clamp to burst range bounds
    lower_isi = max(lower_isi, lo_ms)
    upper_isi = min(upper_isi, hi_ms)

    result.update({
        'peak_isi_ms':   round(peak_isi,   3),
        'peak_freq_hz':  round(1000.0 / peak_isi,   2),
        'lower_isi_ms':  round(lower_isi,  3),
        'upper_isi_ms':  round(upper_isi,  3),
        'lower_freq_hz': round(1000.0 / upper_isi,  2),  # short ISI = high freq
        'upper_freq_hz': round(1000.0 / lower_isi,  2),
    })
    return result


# ── Per-unit ISI plot ─────────────────────────────────────────────────────────

def plot_isi(isis_ms: np.ndarray, burst_result: dict, title: str,
             csi_og: float = float('nan'),
             bins_ms: np.ndarray = ISI_HIST_BINS_MS):
    """ISI histogram with burst range highlighted and peak/limit lines."""
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor('#F7F9FC')
    ax.set_facecolor('#F7F9FC')

    valid = isis_ms[(isis_ms > 0) & (isis_ms < bins_ms[-1])]
    if len(valid):
        ax.hist(valid, bins=bins_ms.tolist(), color='#4A90D9', edgecolor='white',
                linewidth=0.4, zorder=3)

    # Burst range shading
    ax.axvspan(ISI_BURST_LO_MS, ISI_BURST_HI_MS, alpha=0.12, color='#E06C75', zorder=2)
    ax.axvline(ISI_BURST_LO_MS, color='#E06C75', linewidth=1.2, linestyle='--', zorder=4,
               label=f'Burst range [{ISI_BURST_LO_MS}–{ISI_BURST_HI_MS} ms]')
    ax.axvline(ISI_BURST_HI_MS, color='#E06C75', linewidth=1.2, linestyle='--', zorder=4)

    if burst_result.get('has_burst_peak'):
        p   = burst_result['peak_isi_ms']
        lo  = burst_result['lower_isi_ms']
        hi  = burst_result['upper_isi_ms']
        pf  = burst_result['peak_freq_hz']
        lof = burst_result['upper_freq_hz']   # upper_freq ↔ lower_isi
        hif = burst_result['lower_freq_hz']
        ax.axvline(p,  color='#E5C07B', linewidth=2.0, zorder=5,
                   label=f'Peak {p:.1f} ms  ({pf:.1f} Hz)')
        ax.axvline(lo, color='#98C379', linewidth=1.5, linestyle=':', zorder=5,
                   label=f'Lower HM {lo:.1f} ms  ({lof:.1f} Hz)')
        ax.axvline(hi, color='#C678DD', linewidth=1.5, linestyle=':', zorder=5,
                   label=f'Upper HM {hi:.1f} ms  ({hif:.1f} Hz)')

    pct         = burst_result.get('burst_power_pct', float('nan'))
    sig         = burst_result.get('has_burst_peak', False)
    pct_str     = f'{pct:.1f}%' if not (isinstance(pct, float) and np.isnan(pct)) else 'N/A'
    csi_og_str  = f'{csi_og:.2f}%' if not (isinstance(csi_og, float) and np.isnan(csi_og)) else 'N/A'
    ax.text(0.98, 0.97,
            f'CSI_OG: {csi_og_str}  (threshold {CSI_THRESHOLD}%)\n'
            f'Burst power: {pct_str}\n'
            f'{"SIGNIFICANT" if sig else "not significant"}',
            transform=ax.transAxes, fontsize=9, va='top', ha='right',
            bbox=dict(boxstyle='round,pad=0.4', facecolor='white',
                      edgecolor='#CCCCCC', alpha=0.9))

    ax.set_xlim(0, 80)
    ax.set_xlabel('ISI (ms)', fontsize=11, labelpad=8)
    ax.set_ylabel('Spike count', fontsize=11, labelpad=8)
    ax.set_title(title, fontsize=12, fontweight='bold', pad=10)
    ax.legend(fontsize=8.5, loc='upper right', framealpha=0.85, edgecolor='#CCCCCC')
    ax.spines[['top', 'right']].set_visible(False)
    ax.yaxis.grid(True, color='white', linewidth=0.8, zorder=2)
    plt.tight_layout()
    return fig


# ── Batch scan ────────────────────────────────────────────────────────────────

_BURST_NAN = {
    'burst_power_pct': float('nan'), 'has_burst_peak': False,
    'peak_isi_ms': float('nan'),     'peak_freq_hz':  float('nan'),
    'lower_isi_ms': float('nan'),    'upper_isi_ms':  float('nan'),
    'lower_freq_hz': float('nan'),   'upper_freq_hz': float('nan'),
}

_SUA_NAN = {
    'is_sua': False, 'rpv_rate': float('nan'),
    'waveform_snr': float('nan'), 'waveform_cv': float('nan'),
}

records       = []
isi_plot_data = []   # (title, isis_ms, csi, burst_result) for each unit
sua_files     = []   # (src_path, dst_path) for SUA units

for dirpath, dirnames, filenames in os.walk(ROOT_FOLDER):
    dirnames[:] = [d for d in dirnames if d != SUA_SUBDIR_NAME]
    ntt_files = sorted(f for f in filenames if f.lower().endswith('.ntt'))
    if not ntt_files:
        continue
    for ntt_file in ntt_files:
        ntt_path     = os.path.join(dirpath, ntt_file)
        session_name = os.path.relpath(dirpath, ROOT_FOLDER)
        print(f'Processing: {session_name}  |  {ntt_file}')
        try:
            waveforms, spike_ts, _ = load_ntt_waveforms(ntt_path)
            n_spikes   = len(spike_ts)
            isis_ms    = (np.diff(np.sort(spike_ts)) / 1000.0
                          if n_spikes >= 2 else np.array([]))

            sua_result = classify_sua_mua(waveforms, spike_ts) if n_spikes >= 2 else dict(_SUA_NAN)
            is_sua     = sua_result['is_sua']

            if is_sua:
                sua_dir  = os.path.join(dirpath, SUA_SUBDIR_NAME)
                dst_path = os.path.join(sua_dir, ntt_file)
                sua_files.append((ntt_path, dst_path))

            # CSI and burst analysis only for SUA units
            csi_og       = compute_csi(waveforms, spike_ts) if (is_sua and n_spikes >= 50) else float('nan')
            is_complex   = (not (isinstance(csi_og, float) and np.isnan(csi_og))
                            and csi_og > CSI_THRESHOLD)
            burst_result = compute_burst_freq_range(spike_ts) if is_complex else dict(_BURST_NAN)

        except Exception as exc:
            print(f'  ERROR: {exc}')
            n_spikes     = 0
            sua_result   = dict(_SUA_NAN)
            is_sua       = False
            csi_og       = float('nan')
            is_complex   = False
            burst_result = dict(_BURST_NAN)
            isis_ms      = np.array([])

        row = {
            'session': session_name, 'unit': ntt_file, 'n_spikes': n_spikes,
            'is_sua': is_sua,
            'rpv_rate': sua_result['rpv_rate'],
            'waveform_snr': sua_result['waveform_snr'],
            'waveform_cv': sua_result['waveform_cv'],
            'csi_og': csi_og, 'is_complex_cell': is_complex,
        }
        row.update(burst_result)
        records.append(row)
        isi_plot_data.append((f'{session_name} | {ntt_file}', isis_ms, csi_og, burst_result))
        sua_str     = 'SUA' if is_sua else 'MUA'
        rpv_str     = f'{sua_result["rpv_rate"]:.2f}%' if not np.isnan(sua_result['rpv_rate']) else 'N/A'
        snr_str     = f'{sua_result["waveform_snr"]:.2f}' if not np.isnan(sua_result['waveform_snr']) else 'N/A'
        cv_str      = f'{sua_result["waveform_cv"]:.2f}'  if not np.isnan(sua_result['waveform_cv'])  else 'N/A'
        csi_og_str  = f'CSI_OG={csi_og:.2f}%  complex={is_complex}' if not (isinstance(csi_og, float) and np.isnan(csi_og)) else 'CSI_OG=N/A (MUA — skipped)'
        print(f'  {sua_str}  RPV={rpv_str}  SNR={snr_str}  CV={cv_str}  {csi_og_str}')

# ── Build DataFrame ───────────────────────────────────────────────────────────

df = pd.DataFrame(records, columns=COLUMN_ORDER)
df.to_excel(OUTPUT_EXCEL, index=False, engine='openpyxl')

# ── Copy SUA .ntt files into a per-folder SUA subdirectory ────────────────────
if sua_files:
    for src_path, dst_path in sua_files:
        os.makedirs(os.path.dirname(dst_path), exist_ok=True)
        shutil.copy2(src_path, dst_path)
        print(f'  Copied SUA: {dst_path}')
    print(f'Copied {len(sua_files)} SUA file(s) into their respective "{SUA_SUBDIR_NAME}" subfolders')

# ── Per-unit ISI plots ────────────────────────────────────────────────────────

base = os.path.splitext(OUTPUT_EXCEL)[0]
unit_png_paths = []   # (title, png_path, burst_result) for embedding

for title, isis_ms, csi, burst_result in isi_plot_data:
    safe_name = title.replace(os.sep, '_').replace(' ', '_').replace('|', '').strip('_')
    png_path  = f'{base}_ISI_{safe_name}.png'
    fig = plot_isi(isis_ms, burst_result, title, csi_og=csi) # type: ignore
    fig.savefig(png_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    unit_png_paths.append((title, png_path, burst_result))
    print(f'  ISI plot: {os.path.basename(png_path)}')

# ── Post-process Excel: colour coding ────────────────────────────────────────

wb = load_workbook(OUTPUT_EXCEL)
ws = wb.active
assert ws is not None
ws.title = 'Burst Freq Range'

COL_WIDTHS = {
    'session':          28, 'unit':            18,
    'n_spikes':         10,
    'is_sua':           10, 'rpv_rate':        12,
    'waveform_snr':     14, 'waveform_cv':     14,
    'csi_og':           10, 'is_complex_cell': 16,
    'burst_power_pct':  18, 'has_burst_peak':  16,
    'peak_isi_ms':      14, 'peak_freq_hz':    14,
    'lower_isi_ms':     14, 'upper_isi_ms':    14,
    'lower_freq_hz':    14, 'upper_freq_hz':   14,
}
for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 14)

# Header row
for cell in ws[1]:
    cell.fill = FILL_HEADER; cell.font = FONT_HEADER
    cell.alignment = ALIGN_CTR; cell.border = THIN_BORDER
ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

sua_col      = COLUMN_ORDER.index('is_sua')          + 1
rpv_col      = COLUMN_ORDER.index('rpv_rate')         + 1
snr_col      = COLUMN_ORDER.index('waveform_snr')     + 1
cv_col       = COLUMN_ORDER.index('waveform_cv')      + 1
csi_col      = COLUMN_ORDER.index('csi_og')            + 1
complex_col  = COLUMN_ORDER.index('is_complex_cell')  + 1
pwr_col      = COLUMN_ORDER.index('burst_power_pct')  + 1
flag_col     = COLUMN_ORDER.index('has_burst_peak')   + 1

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    for cell in row:
        cell.font      = FONT_BODY
        cell.border    = THIN_BORDER
        cell.alignment = ALIGN_CTR if cell.column > 1 else ALIGN_LEFT  # type: ignore[operator]

    # ── SUA / MUA colour (green = SUA, red = MUA) ────────────────────────────
    is_sua_val = ws.cell(row=row_idx, column=sua_col).value
    sua_fill   = FILL_GREEN if is_sua_val else FILL_RED
    for col_idx in (sua_col, rpv_col, snr_col, cv_col):
        ws.cell(row=row_idx, column=col_idx).fill = sua_fill

    # ── CSI colour ────────────────────────────────────────────────────────────
    try:
        csi_val = float(ws.cell(row=row_idx, column=csi_col).value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        csi_val = float('nan')

    if np.isnan(csi_val):
        csi_fill = PatternFill()
    elif csi_val > CSI_THRESHOLD:
        csi_fill = FILL_GREEN
    elif csi_val > CSI_THRESHOLD / 2:
        csi_fill = FILL_YELLOW
    else:
        csi_fill = FILL_RED

    ws.cell(row=row_idx, column=csi_col).fill     = csi_fill
    ws.cell(row=row_idx, column=complex_col).fill = csi_fill

    # ── Burst power colour (only meaningful for complex cells) ────────────────
    try:
        pwr = float(ws.cell(row=row_idx, column=pwr_col).value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        pwr = float('nan')

    try:
        is_sig = bool(ws.cell(row=row_idx, column=flag_col).value)
    except (TypeError, ValueError):
        is_sig = False

    if is_sig:
        fill = FILL_GREEN if pwr >= 10.0 else FILL_YELLOW
    elif not np.isnan(pwr):
        fill = FILL_RED
    else:
        fill = PatternFill()

    ws.cell(row=row_idx, column=pwr_col).fill  = fill
    ws.cell(row=row_idx, column=flag_col).fill = fill

wb.save(OUTPUT_EXCEL)

# ── Summary histograms ────────────────────────────────────────────────────────

complex_df = df[df['is_complex_cell'] == True]
sig_df     = df[df['has_burst_peak']  == True]

summary_specs = [
    ('Hist_CSI_OG',      '_hist_CSI_OG',      'csi_og',
     'Complex Spike Index',
     'CSI_OG (%)',
     (CSI_THRESHOLD,), ('#333333',), (f'Threshold {CSI_THRESHOLD}%',), '#4472C4'),

    ('Hist_BurstPower',  '_hist_BurstPower',  'burst_power_pct',
     f'ISI Burst Power (SUA complex cells, n={len(complex_df)})',
     f'% of ISIs in [{ISI_BURST_LO_MS}–{ISI_BURST_HI_MS} ms]',
     (SIGNIFICANCE_PCT,), ('#333333',), (f'Threshold {SIGNIFICANCE_PCT}%',), '#ED7D31'),

    ('Hist_PeakISI',     '_hist_PeakISI',     'peak_isi_ms',
     f'Peak ISI (complex cells with burst peak, n={len(sig_df)})',
     'Peak ISI (ms)',
     (), (), (), '#70AD47'),

    ('Hist_PeakFreq',    '_hist_PeakFreq',    'peak_freq_hz',
     f'Peak Burst Frequency (complex cells with burst peak, n={len(sig_df)})',
     'Peak frequency (Hz)',
     (), (), (), '#C00000'),
]

png_summary = {}
for sheet_name, suffix, col, title, xlabel, thresholds, thr_colors, thr_labels, bar_color in summary_specs:
    if 'Peak' in sheet_name:
        source = sig_df
    elif sheet_name == 'Hist_BurstPower':
        source = complex_df
    else:
        source = df
    vals = source[col].dropna().values if col in source.columns else np.array([])

    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')

    if len(vals):
        ax.hist(vals, bins=30, color=bar_color, edgecolor='#555555', linewidth=0.3, zorder=3)
        for xv, col_c, lbl in zip(thresholds, thr_colors, thr_labels):
            ax.axvline(xv, color=col_c, linewidth=1.5, linestyle='--', zorder=4, label=lbl)
        mean_v   = float(np.mean(vals))
        median_v = float(np.median(vals))
        ax.axvline(mean_v,   color='#C00000', linewidth=1.8, linestyle='-',  zorder=5,
                   label=f'Mean   {mean_v:.3f}')
        ax.axvline(median_v, color='#264478', linewidth=1.8, linestyle=':', zorder=5,
                   label=f'Median {median_v:.3f}')
        ax.legend(fontsize=8.5, framealpha=0.85, loc='upper right', edgecolor='#CCCCCC')
    else:
        ax.text(0.5, 0.5, 'No valid data', transform=ax.transAxes,
                ha='center', va='center', fontsize=12)

    ax.set_xlabel(xlabel, fontsize=11, labelpad=8)
    ax.set_ylabel('Number of units', fontsize=11, labelpad=8)
    ax.set_title(title, fontsize=13, fontweight='bold', pad=12)
    ax.spines[['top', 'right']].set_visible(False)
    plt.tight_layout()

    png = base + suffix + '.png'
    svg = base + suffix + '.svg'
    fig.savefig(png, dpi=300, bbox_inches='tight')
    fig.savefig(svg, bbox_inches='tight')
    plt.close(fig)
    png_summary[sheet_name] = png
    print(f'Summary histogram saved: {os.path.basename(png)}  +  {os.path.basename(svg)}')

# ── Embed summary histograms in Excel ─────────────────────────────────────────

wb2 = load_workbook(OUTPUT_EXCEL)

for sheet_name, suffix, col, title, *_ in summary_specs:
    hs = wb2.create_sheet(sheet_name)
    hs.sheet_view.showGridLines = False
    tc = hs.cell(row=1, column=1, value=title)
    tc.font      = Font(bold=True, size=13, color='FF2E4057', name='Calibri')
    tc.alignment = Alignment(horizontal='left', vertical='center')
    hs.row_dimensions[1].height = 24
    img = XLImage(png_summary[sheet_name])
    img.anchor = 'A3'
    hs.add_image(img)

# Embed per-unit ISI plots as individual sheets (one per significant unit)
for title, png_path, burst_result in unit_png_paths:
    if not burst_result.get('has_burst_peak'):
        continue
    # Excel forbids \ / ? * [ ] : in sheet names; limit is 31 chars
    _illegal = str.maketrans({c: '_' for c in r'\/?*[]:'})
    sheet_label = title.replace(' | ', '_').translate(_illegal)[:28]
    hs = wb2.create_sheet(sheet_label)
    hs.sheet_view.showGridLines = False
    tc = hs.cell(row=1, column=1, value=title)
    tc.font      = Font(bold=True, size=11, color='FF2E4057', name='Calibri')
    tc.alignment = Alignment(horizontal='left', vertical='center')
    hs.row_dimensions[1].height = 22
    img = XLImage(png_path)
    img.anchor = 'A3'
    hs.add_image(img)

wb2.save(OUTPUT_EXCEL)

# ── Console summary ───────────────────────────────────────────────────────────

sua_df = df[df['is_sua'] == True]

print(f'\nDone. Results saved to {OUTPUT_EXCEL}')
print(f'Total units processed : {len(df)}')
print(f'  SUA (RPV<{RPV_MAX_PCT}%, SNR>={WF_SNR_MIN}, CV<={WF_CV_MAX}): {len(sua_df)}')
print(f'  MUA                  : {len(df) - len(sua_df)}')
print(f'Complex cells (CSI > {CSI_THRESHOLD}%, SUA only): {int(df["is_complex_cell"].sum())}')
print(f'Burst peak found      : {int(df["has_burst_peak"].sum())} '
      f'(complex cell + burst_power_pct >= {SIGNIFICANCE_PCT}%)')

if len(sig_df):
    for col, label in [('peak_isi_ms', 'Peak ISI (ms)'),
                        ('peak_freq_hz', 'Peak freq (Hz)'),
                        ('lower_isi_ms', 'Lower HM ISI (ms)'),
                        ('upper_isi_ms', 'Upper HM ISI (ms)'),
                        ('lower_freq_hz', 'Lower freq (Hz)'),
                        ('upper_freq_hz', 'Upper freq (Hz)')]:
        v = sig_df[col].dropna()
        if len(v):
            print(f'  {label:22s}: mean={v.mean():.3f}  '
                  f'median={v.median():.3f}  SD={v.std():.3f}  n={len(v)}')
